from __future__ import annotations

import base64
import csv
import io
import json
import math
import mimetypes
import os
import re
import sqlite3
from collections import defaultdict, deque
from datetime import date, datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    load_workbook = None
    DataBarRule = None
    FormulaRule = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None
    get_column_letter = None


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
DATA_DIR = Path(os.environ.get("TASK_GANTT_DATA_DIR", str(ROOT / "data"))).expanduser()
DB_PATH = DATA_DIR / "task_gantt.db"
DATE_FMT = "%Y-%m-%d"
HOURS_PER_DAY = 6.0
ENV_PATH = ROOT / ".env"
SILICONFLOW_DEFAULT_URL = "https://api.siliconflow.cn/v1/chat/completions"
SILICONFLOW_DEFAULT_MODEL = "deepseek-ai/DeepSeek-V3.2"


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if len(value) >= 2 and value[0] == value[-1] and value[0] in ('"', "'"):
            value = value[1:-1]
        os.environ.setdefault(key, value)


load_env_file(ENV_PATH)


PROJECT_TEMPLATES = {
    "web_tool": [
        {
            "key": "scope",
            "title": "需求梳理与范围冻结",
            "description": "确认目标、角色、核心流程和最小可交付范围。",
            "complexity": 2,
            "priority": 1,
            "depends_on": [],
        },
        {
            "key": "model",
            "title": "数据模型与字段设计",
            "description": "定义项目、任务、依赖、估时记录和导入导出字段。",
            "complexity": 3,
            "priority": 1,
            "depends_on": ["scope"],
        },
        {
            "key": "backend",
            "title": "后端存储与接口实现",
            "description": "实现表存储、任务 CRUD、统计指标与调度接口。",
            "complexity": 4,
            "priority": 1,
            "depends_on": ["model"],
        },
        {
            "key": "frontend",
            "title": "前端管理页与交互",
            "description": "实现项目面板、任务编辑、状态切换和筛选。",
            "complexity": 4,
            "priority": 1,
            "depends_on": ["model"],
        },
        {
            "key": "gantt",
            "title": "甘特图视图与排期逻辑",
            "description": "实现时间轴渲染、依赖关系展示和拖动改期。",
            "complexity": 4,
            "priority": 2,
            "depends_on": ["backend", "frontend"],
        },
        {
            "key": "io",
            "title": "导入导出能力",
            "description": "支持 CSV、JSON、XLSX 的导入和导出。",
            "complexity": 3,
            "priority": 2,
            "depends_on": ["backend"],
        },
        {
            "key": "smart",
            "title": "项目识别与估时规则",
            "description": "根据项目描述自动拆解任务，并给出估时与置信度。",
            "complexity": 4,
            "priority": 2,
            "depends_on": ["backend"],
        },
        {
            "key": "qa",
            "title": "联调测试与样例校准",
            "description": "校验排期逻辑、导入导出结果和估时建议。",
            "complexity": 3,
            "priority": 2,
            "depends_on": ["gantt", "io", "smart"],
        },
        {
            "key": "release",
            "title": "文档整理与发布验收",
            "description": "补齐使用说明、示例数据和验收清单。",
            "complexity": 2,
            "priority": 3,
            "depends_on": ["qa"],
        },
    ],
    "data_project": [
        {
            "key": "scope",
            "title": "业务问题拆解",
            "description": "明确输入数据、指标口径和输出格式。",
            "complexity": 2,
            "priority": 1,
            "depends_on": [],
        },
        {
            "key": "source",
            "title": "数据源梳理与清洗规则",
            "description": "明确采集来源、字段映射和清洗异常策略。",
            "complexity": 3,
            "priority": 1,
            "depends_on": ["scope"],
        },
        {
            "key": "pipeline",
            "title": "处理流程与脚本实现",
            "description": "实现数据处理、转换和结果产出流程。",
            "complexity": 4,
            "priority": 1,
            "depends_on": ["source"],
        },
        {
            "key": "analysis",
            "title": "指标分析与验证",
            "description": "对核心结果进行对照、采样和异常回看。",
            "complexity": 4,
            "priority": 2,
            "depends_on": ["pipeline"],
        },
        {
            "key": "report",
            "title": "报告、图表与交付说明",
            "description": "整理结论、可视化图表和业务说明。",
            "complexity": 3,
            "priority": 2,
            "depends_on": ["analysis"],
        },
    ],
    "content_project": [
        {
            "key": "scope",
            "title": "主题定位与受众定义",
            "description": "确定目标受众、输出调性和关键约束。",
            "complexity": 2,
            "priority": 1,
            "depends_on": [],
        },
        {
            "key": "outline",
            "title": "结构大纲与素材清单",
            "description": "拆解章节结构、素材类型和引用来源。",
            "complexity": 3,
            "priority": 1,
            "depends_on": ["scope"],
        },
        {
            "key": "draft",
            "title": "初稿产出",
            "description": "完成核心内容撰写或创作初版。",
            "complexity": 4,
            "priority": 1,
            "depends_on": ["outline"],
        },
        {
            "key": "review",
            "title": "审校与迭代",
            "description": "处理反馈、统一风格并修正逻辑问题。",
            "complexity": 3,
            "priority": 2,
            "depends_on": ["draft"],
        },
        {
            "key": "delivery",
            "title": "交付封装",
            "description": "整理最终文件、封面、目录和版本说明。",
            "complexity": 2,
            "priority": 2,
            "depends_on": ["review"],
        },
    ],
    "general": [
        {
            "key": "scope",
            "title": "范围澄清",
            "description": "明确目标、完成标准和关键约束。",
            "complexity": 2,
            "priority": 1,
            "depends_on": [],
        },
        {
            "key": "breakdown",
            "title": "任务拆分",
            "description": "将目标拆为可执行子任务和里程碑。",
            "complexity": 3,
            "priority": 1,
            "depends_on": ["scope"],
        },
        {
            "key": "execute",
            "title": "核心执行",
            "description": "完成主要交付内容与过程协调。",
            "complexity": 4,
            "priority": 1,
            "depends_on": ["breakdown"],
        },
        {
            "key": "review",
            "title": "复核与修订",
            "description": "检查结果质量并完成修订。",
            "complexity": 3,
            "priority": 2,
            "depends_on": ["execute"],
        },
        {
            "key": "delivery",
            "title": "交付收尾",
            "description": "整理输出、归档资料和验收。",
            "complexity": 2,
            "priority": 2,
            "depends_on": ["review"],
        },
    ],
}


TASK_RULES = [
    (("research", "调研", "梳理", "scope", "需求"), 6.0),
    (("model", "schema", "字段", "数据", "database", "存储"), 10.0),
    (("backend", "接口", "api", "服务"), 14.0),
    (("frontend", "页面", "界面", "web", "ui"), 12.0),
    (("gantt", "排期", "甘特"), 16.0),
    (("import", "export", "导入", "导出", "xlsx", "csv"), 8.0),
    (("estimate", "预估", "识别", "智能", "规则"), 10.0),
    (("test", "qa", "联调", "验收"), 8.0),
    (("deploy", "release", "上线", "发布"), 5.0),
    (("doc", "文档", "说明"), 4.0),
    (("analysis", "分析", "pipeline", "脚本"), 12.0),
    (("report", "报告", "图表"), 8.0),
]


def today_str() -> str:
    return date.today().strftime(DATE_FMT)


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.strptime(str(value)[:10], DATE_FMT).date()


def format_date(value: date | None) -> str | None:
    return value.strftime(DATE_FMT) if value else None


def add_days(value: str | None, days: int) -> str:
    current = parse_date(value) or date.today()
    return format_date(current + timedelta(days=days))


def clamp(number: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, number))


def parse_float(value, default: float = 0.0) -> float:
    try:
        if value in ("", None):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_int(value, default: int = 0) -> int:
    try:
        if value in ("", None):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def parse_bool(value, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def infer_category(name: str, description: str) -> str:
    text = f"{name} {description}".lower()
    if any(token in text for token in ("网页", "web", "前端", "backend", "api", "工具", "dashboard", "gantt", "甘特")):
        return "web_tool"
    if any(token in text for token in ("data", "数据", "分析", "pipeline", "etl", "报表")):
        return "data_project"
    if any(token in text for token in ("content", "文章", "文档", "课程", "内容", "脚本", "视频")):
        return "content_project"
    return "general"


def estimate_task_hours(
    title: str,
    description: str,
    complexity: int,
    dependency_count: int,
    category: str,
) -> tuple[float, float, str]:
    text = f"{title} {description}".lower()
    base_hours = 6.0
    reason = "通用基线"
    for keywords, hours in TASK_RULES:
        if any(keyword in text for keyword in keywords):
            base_hours = max(base_hours, hours)
            reason = f"关键词匹配: {keywords[0]}"
    if category == "web_tool":
        base_hours += 1.5
    elif category == "data_project":
        base_hours += 1.0
    hours = base_hours * (0.75 + complexity * 0.35) + dependency_count * 1.5
    confidence = 0.56
    if reason != "通用基线":
        confidence += 0.12
    confidence += (3 - abs(complexity - 3)) * 0.04
    confidence -= max(0, dependency_count - 1) * 0.05
    return round(hours, 1), round(clamp(confidence, 0.45, 0.92), 2), reason


def apply_due_date_pressure(tasks: list[dict], start_date: str, due_date: str | None) -> None:
    if not due_date:
        return
    project_start = parse_date(start_date) or date.today()
    project_due = parse_date(due_date)
    if project_due and project_due < project_start + timedelta(days=len(tasks)):
        for task in tasks:
            task["estimate_hours"] = round(task["estimate_hours"] * 0.9, 1)
            task["estimate_basis"] += " / 截止日期偏紧"


def schedule_task_batch(tasks: list[dict], start_date: str) -> None:
    if not tasks:
        return
    batch_start = parse_date(start_date) or date.today()
    ordered_keys = [str(task.get("key") or f"task_{index + 1}") for index, task in enumerate(tasks)]
    task_map = {ordered_keys[index]: tasks[index] for index in range(len(tasks))}
    indegree = {key: 0 for key in ordered_keys}
    graph: dict[str, list[str]] = defaultdict(list)
    reverse_dependencies: dict[str, list[str]] = defaultdict(list)

    for key in ordered_keys:
        dependency_keys = []
        for dependency in ensure_list(task_map[key].get("dependency_keys")):
            dependency_key = str(dependency).strip()
            if dependency_key in task_map and dependency_key != key and dependency_key not in dependency_keys:
                dependency_keys.append(dependency_key)
        task_map[key]["dependency_keys"] = dependency_keys
        for dependency_key in dependency_keys:
            graph[dependency_key].append(key)
            reverse_dependencies[key].append(dependency_key)
            indegree[key] += 1

    order: list[str] = []
    queue = deque(key for key in ordered_keys if indegree[key] == 0)
    while queue:
        current = queue.popleft()
        order.append(current)
        for neighbor in graph.get(current, []):
            indegree[neighbor] -= 1
            if indegree[neighbor] == 0:
                queue.append(neighbor)

    if len(order) != len(ordered_keys):
        order = ordered_keys

    schedule_end: dict[str, date] = {}
    for key in order:
        dependency_end_dates = [schedule_end[item] for item in reverse_dependencies.get(key, []) if item in schedule_end]
        task_start = max([batch_start, *dependency_end_dates], default=batch_start)
        if dependency_end_dates:
            task_start += timedelta(days=1)
        duration_days = task_duration_days(task_map[key].get("estimate_hours", 6.0))
        task_end = task_start + timedelta(days=duration_days - 1)
        task_map[key]["start_date"] = format_date(task_start)
        task_map[key]["end_date"] = format_date(task_end)
        schedule_end[key] = task_end


def sanitize_task_key(raw_value: str, index: int, used_keys: set[str]) -> str:
    base = re.sub(r"[^a-z0-9]+", "_", str(raw_value or "").strip().lower()).strip("_")
    if not base:
        base = f"task_{index}"
    if base[0].isdigit():
        base = f"task_{base}"
    candidate = base
    suffix = 2
    while candidate in used_keys:
        candidate = f"{base}_{suffix}"
        suffix += 1
    used_keys.add(candidate)
    return candidate


def ensure_list(value) -> list:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    text = str(value).replace("，", ",").replace("、", ",")
    return [part.strip() for part in text.split(",") if part.strip()]


def extract_json_object(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    decoder = json.JSONDecoder()
    for index, char in enumerate(cleaned):
        if char != "{":
            continue
        try:
            payload, _ = decoder.raw_decode(cleaned[index:])
            if isinstance(payload, dict):
                return payload
        except json.JSONDecodeError:
            continue
    raise ValueError("LLM response is not a valid JSON object")


MERMAID_GANTT_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
MERMAID_GANTT_DURATION_RE = re.compile(r"^(?P<days>\d+)d$", re.IGNORECASE)
MERMAID_GANTT_STATUS_MAP = {
    "done": ("done", 100),
    "active": ("in_progress", 45),
    "crit": ("blocked", 20),
    "milestone": ("done", 100),
}


OUTLINE_BULLET_CHAR = chr(0x2022)
OUTLINE_FULLWIDTH_COLON = chr(0xFF1A)
OUTLINE_STAGE_WORD = chr(0x9636) + chr(0x6BB5)

OUTLINE_STAGE_RE = re.compile(
    rf"^\s*(?:[{re.escape(OUTLINE_BULLET_CHAR)}*-]\s*)?(?:{OUTLINE_STAGE_WORD}|phase)\s*(?P<num>\d+)\s*[:{OUTLINE_FULLWIDTH_COLON}]\s*(?P<title>.+?)\s*$",
    re.IGNORECASE,
)
OUTLINE_NUMBERED_RE = re.compile(
    rf"^\s*(?:[{re.escape(OUTLINE_BULLET_CHAR)}*-]\s*)?(?P<num>\d+(?:\.\d+)+)\s*(?P<body>.+?)\s*$"
)
OUTLINE_BULLET_RE = re.compile(rf"^\s*[{re.escape(OUTLINE_BULLET_CHAR)}*-]\s*(?P<body>.+?)\s*$")


def contains_cjk(text: str) -> bool:
    return any("一" <= char <= "鿿" for char in str(text or ""))


def looks_like_mermaid_gantt(text: str) -> bool:
    if not text:
        return False
    lowered = text.lower()
    return "gantt" in lowered and "section" in lowered and "title" in lowered and ":" in text


def looks_like_numbered_outline(text: str) -> bool:
    if not text:
        return False
    lines = [line.strip() for line in str(text).splitlines() if line.strip()]
    has_stage = any(OUTLINE_STAGE_RE.match(line) for line in lines)
    numbered_count = 0
    for line in lines:
        if OUTLINE_NUMBERED_RE.match(line):
            numbered_count += 1
            continue
        bullet_match = OUTLINE_BULLET_RE.match(line)
        if bullet_match and OUTLINE_NUMBERED_RE.match(bullet_match.group("body").strip()):
            numbered_count += 1
    return numbered_count >= 2 and (has_stage or numbered_count >= 3)


def split_outline_item(text: str) -> tuple[str, str]:
    cleaned = re.sub(r"\s+", " ", str(text or "")).strip().strip(OUTLINE_BULLET_CHAR + "*- ")
    for separator in (OUTLINE_FULLWIDTH_COLON, ":"):
        if separator in cleaned:
            title, detail = cleaned.split(separator, 1)
            return title.strip(), detail.strip()
    return cleaned, ""


def infer_priority_from_section(section: str, status: str) -> int:
    match = re.search(r"\bP(\d+)\b", section, re.IGNORECASE)
    if match:
        return max(1, min(5, int(match.group(1))))
    if status == "done":
        return 1
    return 2


def infer_complexity_from_duration(duration_days: int) -> int:
    if duration_days >= 14:
        return 5
    if duration_days >= 10:
        return 4
    if duration_days >= 6:
        return 3
    if duration_days >= 3:
        return 2
    return 1


def merge_llm_tasks_with_structured_hint(llm_tasks: list[dict], structured_tasks: list[dict]) -> list[dict]:
    if not structured_tasks:
        return llm_tasks
    merged_tasks: list[dict] = []
    for index, hint_task in enumerate(structured_tasks):
        llm_task = llm_tasks[index] if index < len(llm_tasks) else {}
        merged_task = dict(hint_task)
        if llm_task:
            merged_task["description"] = llm_task.get("description") or merged_task.get("description", "")
            merged_task["priority"] = max(1, min(5, parse_int(llm_task.get("priority"), merged_task.get("priority", 2))))
            merged_task["complexity"] = max(1, min(5, parse_int(llm_task.get("complexity"), merged_task.get("complexity", 3))))
            merged_task["estimate_hours"] = round(
                max(0.5, parse_float(llm_task.get("estimate_hours"), merged_task.get("estimate_hours", 6.0))),
                1,
            )
            merged_task["confidence"] = round(
                clamp(parse_float(llm_task.get("confidence"), merged_task.get("confidence", 0.9)), 0.45, 0.95),
                2,
            )
            merged_task["estimate_basis"] = llm_task.get("estimate_basis") or merged_task.get("estimate_basis", "")
            note_parts = [part for part in (merged_task.get("notes", ""), llm_task.get("notes", "")) if part]
            merged_task["notes"] = "\uFF1B".join(dict.fromkeys(note_parts))
        merged_tasks.append(merged_task)
    return merged_tasks


def parse_mermaid_gantt_tasks(
    name: str,
    description: str,
    fallback_start_date: str,
    fallback_due_date: str | None = None,
) -> dict | None:
    if not looks_like_mermaid_gantt(description):
        return None

    lines = [line.rstrip() for line in description.splitlines() if line.strip()]
    fallback_start = parse_date(fallback_start_date) or date.today()
    fallback_due = parse_date(fallback_due_date)
    parsed_title = ""
    current_section = ""
    used_keys: set[str] = set()
    alias_map: dict[str, str] = {}
    records: list[dict] = []

    for index, raw_line in enumerate(lines, start=1):
        stripped = raw_line.strip()
        lowered = stripped.lower()
        if lowered == "gantt":
            continue
        if lowered.startswith("title "):
            parsed_title = stripped[6:].strip()
            continue
        if lowered.startswith("section "):
            current_section = stripped[8:].strip()
            continue
        if lowered.startswith("dateformat ") or lowered.startswith("axisformat "):
            continue
        if ":" not in stripped:
            continue

        title_part, meta_part = stripped.rsplit(":", 1)
        title = title_part.strip()
        meta_tokens = [token.strip() for token in meta_part.split(",") if token.strip()]
        if not title or not meta_tokens:
            continue

        status_token = ""
        if meta_tokens and meta_tokens[0].lower() in MERMAID_GANTT_STATUS_MAP:
            status_token = meta_tokens.pop(0).lower()
        raw_key = meta_tokens.pop(0) if meta_tokens else title
        key = sanitize_task_key(raw_key, index, used_keys)
        alias_map[key.lower()] = key
        alias_map[str(raw_key).strip().lower()] = key
        alias_map[title.lower()] = key

        raw_dependency_refs: list[str] = []
        explicit_start: str | None = None
        explicit_end: str | None = None
        duration_days: int | None = None

        for token in meta_tokens:
            lowered_token = token.lower()
            if lowered_token.startswith("after "):
                ref_text = token[6:].strip()
                raw_dependency_refs.extend(part.strip() for part in re.split(r"[\s,]+", ref_text) if part.strip())
                continue
            if MERMAID_GANTT_DATE_RE.match(token):
                if explicit_start is None:
                    explicit_start = token
                elif explicit_end is None:
                    explicit_end = token
                continue
            duration_match = MERMAID_GANTT_DURATION_RE.match(lowered_token)
            if duration_match:
                duration_days = max(1, int(duration_match.group("days")))

        status, progress = MERMAID_GANTT_STATUS_MAP.get(status_token, ("planned", 0))
        records.append(
            {
                "key": key,
                "raw_key": str(raw_key),
                "title": title,
                "section": current_section,
                "status": status,
                "progress": progress,
                "explicit_start": explicit_start,
                "explicit_end": explicit_end,
                "duration_days": duration_days,
                "raw_dependency_refs": raw_dependency_refs,
                "dependency_keys": [],
            }
        )

    if not records:
        return None

    by_key = {record["key"]: record for record in records}

    def resolve_schedule(record: dict, visiting: set[str] | None = None) -> tuple[date, date]:
        if record.get("_start_date") and record.get("_end_date"):
            return record["_start_date"], record["_end_date"]
        if visiting is None:
            visiting = set()
        if record["key"] in visiting:
            start_value = parse_date(record.get("explicit_start")) or fallback_start
            end_value = parse_date(record.get("explicit_end")) or start_value
            record["_start_date"] = start_value
            record["_end_date"] = max(start_value, end_value)
            return record["_start_date"], record["_end_date"]

        visiting.add(record["key"])
        dependency_keys: list[str] = []
        dependency_end_dates: list[date] = []
        for raw_ref in record.get("raw_dependency_refs", []):
            target_key = alias_map.get(str(raw_ref).strip().lower())
            if not target_key or target_key == record["key"] or target_key in dependency_keys:
                continue
            dependency_keys.append(target_key)
            target_record = by_key.get(target_key)
            if target_record:
                _, dependency_end = resolve_schedule(target_record, visiting)
                dependency_end_dates.append(dependency_end)
        visiting.discard(record["key"])

        record["dependency_keys"] = dependency_keys
        start_value = parse_date(record.get("explicit_start"))
        if start_value is None and dependency_end_dates:
            start_value = max(dependency_end_dates) + timedelta(days=1)
        if start_value is None:
            start_value = fallback_start

        end_value = parse_date(record.get("explicit_end"))
        if end_value is None:
            duration_days = record.get("duration_days") or 1
            end_value = start_value + timedelta(days=max(1, duration_days) - 1)
        if end_value < start_value:
            end_value = start_value

        record["_start_date"] = start_value
        record["_end_date"] = end_value
        return start_value, end_value

    tasks: list[dict] = []
    start_dates: list[date] = []
    end_dates: list[date] = []
    for record in records:
        start_value, end_value = resolve_schedule(record)
        start_dates.append(start_value)
        end_dates.append(end_value)
        duration_days = max(1, (end_value - start_value).days + 1)
        section_text = record.get("section", "").strip()
        notes = []
        if section_text:
            notes.append(f"阶段：{section_text}")
        if record.get("raw_key"):
            notes.append(f"原始ID：{record['raw_key']}")
        tasks.append(
            {
                "key": record["key"],
                "title": record["title"],
                "description": f"阶段：{section_text}" if section_text else "",
                "status": record["status"],
                "owner": "",
                "priority": infer_priority_from_section(section_text, record["status"]),
                "complexity": infer_complexity_from_duration(duration_days),
                "estimate_hours": round(duration_days * HOURS_PER_DAY, 1),
                "actual_hours": 0.0,
                "progress": record["progress"],
                "start_date": format_date(start_value),
                "end_date": format_date(end_value),
                "dependency_keys": list(record.get("dependency_keys", [])),
                "confidence": 0.9,
                "estimate_basis": "结构化 gantt 文本解析",
                "notes": "\uFF1B".join(notes),
                "auto_generated": 1,
            }
        )

    project_start = min(start_dates) if start_dates else fallback_start
    project_due = max(end_dates) if end_dates else (fallback_due or project_start)
    if fallback_due and fallback_due > project_due:
        project_due = fallback_due

    section_count = len({record.get("section", "").strip() for record in records if record.get("section")})
    done_count = sum(1 for task in tasks if task["status"] == "done")
    summary_title = parsed_title or name
    project_summary = (
        f"{summary_title} \u00B7 \u5DF2\u89E3\u6790 {len(tasks)} \u4E2A\u4EFB\u52A1"
        f"，其中已完成 {done_count} 个，阶段 {section_count} 个。"
    )
    category = infer_category(name, f"{description} {parsed_title}")
    return {
        "category": category,
        "tasks": tasks,
        "source": "structured",
        "model": "",
        "note": f"已按结构化 gantt 文本解析 {len(tasks)} 个任务",
        "project_description": project_summary,
        "start_date": format_date(project_start),
        "due_date": format_date(project_due),
        "preserve_schedule": True,
    }


def parse_numbered_outline_tasks(
    name: str,
    description: str,
    fallback_start_date: str,
    fallback_due_date: str | None = None,
) -> dict | None:
    if not looks_like_numbered_outline(description):
        return None

    lines = [line.strip() for line in description.splitlines() if line.strip()]
    category = infer_category(name, description)
    stage_title = ""
    stage_summaries: dict[str, list[str]] = defaultdict(list)
    current_group = ""
    used_keys: set[str] = set()
    tasks: list[dict] = []
    previous_key = ""
    stage_labels: set[str] = set()

    def append_task(title: str, detail: str, stage_label: str, group_label: str, raw_key: str) -> None:
        nonlocal previous_key
        normalized_title = str(title or "").strip()
        normalized_detail = str(detail or "").strip()
        if not normalized_title:
            return
        dependency_keys = [previous_key] if previous_key else []
        complexity_seed = 3
        complexity_text = f"{normalized_title} {normalized_detail}"
        if any(
            token in complexity_text
            for token in (
                "\u7cfb\u7edf",
                "\u5e73\u53f0",
                "\u96c6\u6210",
                "\u5fae\u8c03",
                "\u514b\u9686",
                "\u6570\u636e\u5e93",
                "Agent",
                "\u6570\u5b57\u4eba",
                "\u591a\u6a21\u6001",
            )
        ):
            complexity_seed = 4
        estimate_hours, confidence, basis = estimate_task_hours(
            normalized_title,
            normalized_detail or stage_label,
            complexity_seed,
            len(dependency_keys),
            category,
        )
        notes: list[str] = []
        if stage_label:
            notes.append(f"\u9636\u6bb5\uff1a{stage_label}")
            stage_labels.add(stage_label)
        if group_label:
            notes.append(f"\u5f52\u5c5e\uff1a{group_label}")
        stage_summary = " ".join(stage_summaries.get(stage_label, []))
        if stage_summary:
            notes.append(f"\u80cc\u666f\uff1a{stage_summary}")
        key = sanitize_task_key(raw_key or normalized_title, len(tasks) + 1, used_keys)
        tasks.append(
            {
                "key": key,
                "title": normalized_title,
                "description": normalized_detail,
                "status": "planned",
                "owner": "",
                "priority": infer_priority_from_section(stage_label, "planned"),
                "complexity": complexity_seed,
                "estimate_hours": estimate_hours,
                "actual_hours": 0.0,
                "progress": 0,
                "start_date": fallback_start_date,
                "end_date": fallback_start_date,
                "dependency_keys": dependency_keys,
                "confidence": confidence,
                "estimate_basis": basis,
                "notes": "\uff1b".join(notes),
                "auto_generated": 1,
            }
        )
        previous_key = key

    for raw_line in lines:
        stage_match = OUTLINE_STAGE_RE.match(raw_line)
        if stage_match:
            stage_title = f"\u9636\u6bb5 {stage_match.group('num')}\uff1a{stage_match.group('title').strip()}"
            current_group = ""
            continue

        numbered_match = OUTLINE_NUMBERED_RE.match(raw_line)
        if numbered_match:
            item_number = numbered_match.group("num")
            body = numbered_match.group("body").strip()
            title, detail = split_outline_item(body)
            if not detail and body.endswith(("\uff1a", ":")):
                current_group = title
                continue
            current_group = ""
            append_task(title, detail, stage_title, "", item_number)
            continue

        bullet_match = OUTLINE_BULLET_RE.match(raw_line)
        if bullet_match and stage_title:
            body = bullet_match.group("body").strip()
            nested_numbered = OUTLINE_NUMBERED_RE.match(body)
            if nested_numbered:
                item_number = nested_numbered.group("num")
                nested_body = nested_numbered.group("body").strip()
                nested_title, nested_detail = split_outline_item(nested_body)
                if not nested_detail and nested_body.endswith(("\uff1a", ":")):
                    current_group = nested_title
                    continue
                append_task(nested_title, nested_detail, stage_title, current_group, item_number)
                continue
            bullet_title, bullet_detail = split_outline_item(body)
            if bullet_title and not bullet_detail and len(bullet_title) <= 30:
                stage_summaries[stage_title].append(bullet_title)
                continue
            append_task(bullet_title, bullet_detail, stage_title, current_group, f"{stage_title}_{len(tasks) + 1}")
            continue

        if stage_title:
            cleaned_line = raw_line.strip("\u2022*- ")
            if cleaned_line:
                stage_summaries[stage_title].append(cleaned_line)

    if not tasks:
        return None

    apply_due_date_pressure(tasks, fallback_start_date, fallback_due_date)
    schedule_task_batch(tasks, fallback_start_date)
    project_due = max(
        (parse_date(task.get("end_date")) for task in tasks if task.get("end_date")),
        default=parse_date(fallback_start_date),
    )
    project_due_str = format_date(project_due) if project_due else fallback_due_date
    stage_count = len(stage_labels)
    project_summary = (
        f"{name} \u00b7 \u5df2\u6309\u9636\u6bb5\u5927\u7eb2\u89e3\u6790 {len(tasks)} "
        f"\u4e2a\u4efb\u52a1\uff0c\u8986\u76d6 {stage_count} \u4e2a\u9636\u6bb5\u3002"
    )
    return {
        "category": category,
        "tasks": tasks,
        "source": "structured",
        "model": "",
        "note": f"\u5df2\u6309\u9636\u6bb5\u5927\u7eb2\u89e3\u6790 {len(tasks)} \u4e2a\u4efb\u52a1",
        "project_description": project_summary,
        "start_date": fallback_start_date,
        "due_date": project_due_str,
        "preserve_schedule": False,
    }

class SiliconFlowPlanner:
    def __init__(self, api_key: str, base_url: str, model: str, timeout_seconds: int = 45):
        self.api_key = api_key.strip()
        self.base_url = (base_url or SILICONFLOW_DEFAULT_URL).strip()
        self.model = (model or SILICONFLOW_DEFAULT_MODEL).strip()
        self.timeout_seconds = max(10, timeout_seconds)

    @classmethod
    def from_env(cls) -> SiliconFlowPlanner | None:
        api_key = os.getenv("SILICONFLOW_API_KEY", "").strip()
        if not api_key:
            return None
        base_url = os.getenv("SILICONFLOW_BASE_URL", SILICONFLOW_DEFAULT_URL).strip()
        model = os.getenv("SILICONFLOW_MODEL", SILICONFLOW_DEFAULT_MODEL).strip()
        timeout_seconds = parse_int(os.getenv("SILICONFLOW_TIMEOUT_SECONDS"), 45)
        return cls(api_key, base_url, model, timeout_seconds)

    def build_task_suggestions(
        self,
        name: str,
        description: str,
        start_date: str,
        due_date: str | None = None,
        structured_hint: dict | None = None,
    ) -> dict:
        input_language = "zh-CN" if contains_cjk(f"{name} {description}") else "same-as-user"
        hint_payload = None
        if structured_hint:
            hint_payload = {
                "source": structured_hint.get("source"),
                "project_description": structured_hint.get("project_description"),
                "start_date": structured_hint.get("start_date"),
                "due_date": structured_hint.get("due_date"),
                "tasks": [
                    {
                        "key": task.get("key"),
                        "title": task.get("title"),
                        "status": task.get("status"),
                        "start_date": task.get("start_date"),
                        "end_date": task.get("end_date"),
                        "dependency_keys": task.get("dependency_keys", []),
                    }
                    for task in structured_hint.get("tasks", [])
                ],
            }

        system_prompt = (
            "You are a project planning assistant. "
            "Return JSON only. Do not return Markdown. "
            "Preserve the user's dominant language. If task titles or section names are Chinese, keep them in Chinese. "
            "Do not translate task titles unless the user explicitly asks for translation. "
            "If structured_hint is provided, treat it as high-confidence schedule data: keep task titles, dates, statuses, "
            "and dependency relationships unless the input clearly contradicts them. "
            "Your job is to normalize the project into this system by adding concise descriptions, priority, complexity, "
            "estimate_hours, confidence, and estimate_basis. "
            "Also produce a short project_description summary suitable for the app overview card."
        )
        payload = {
            "model": self.model,
            "temperature": 0.2,
            "max_tokens": 2200,
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": system_prompt,
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "task": "Normalize the project input into the app task schema.",
                            "language_preference": input_language,
                            "project": {
                                "name": name,
                                "description": description,
                                "start_date": start_date,
                                "due_date": due_date,
                            },
                            "structured_hint": hint_payload,
                            "requirements": {
                                "category_options": ["web_tool", "data_project", "content_project", "general"],
                                "task_count": "Use the natural task count from the input. For vague inputs, prefer 6-10 tasks.",
                                "task_rules": [
                                    "Preserve the user's original language in title and description.",
                                    "If structured_hint is present, keep its task order, title, schedule, status, and dependencies.",
                                    "dependency_keys must reference existing task keys.",
                                    "estimate_hours must be numeric.",
                                    "confidence must be between 0.45 and 0.95.",
                                    "priority and complexity must be integers between 1 and 5.",
                                ],
                                "output_schema": {
                                    "category": "web_tool",
                                    "project_description": "A short one-sentence summary for the project overview.",
                                    "tasks": [
                                        {
                                            "key": "scope",
                                            "title": "Task title in the user's original language",
                                            "description": "Concise execution-oriented description",
                                            "priority": 1,
                                            "complexity": 2,
                                            "estimate_hours": 8,
                                            "confidence": 0.74,
                                            "dependency_keys": [],
                                            "estimate_basis": "Reason for the estimate",
                                            "notes": "",
                                        }
                                    ],
                                },
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        }
        request = Request(
            self.base_url,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.api_key}",
            },
            method="POST",
        )
        with urlopen(request, timeout=self.timeout_seconds) as response:
            response_payload = json.loads(response.read().decode("utf-8"))

        content = ((((response_payload.get("choices") or [{}])[0]).get("message") or {}).get("content") or "").strip()
        if not content:
            raise ValueError("SiliconFlow \u8FD4\u56DE\u7A7A\u54CD\u5E94")
        parsed = extract_json_object(content)
        raw_tasks = parsed.get("tasks") or parsed.get("items") or parsed.get("work_items") or []
        if not isinstance(raw_tasks, list) or not raw_tasks:
            raise ValueError("SiliconFlow \u672A\u8FD4\u56DE\u6709\u6548\u4EFB\u52A1\u5217\u8868")

        category = str(parsed.get("category") or infer_category(name, description)).strip()
        if category not in PROJECT_TEMPLATES:
            category = infer_category(name, description)

        prepared_rows: list[tuple[dict, str, str]] = []
        alias_map: dict[str, str] = {}
        used_keys: set[str] = set()
        for index, raw_task in enumerate(raw_tasks, start=1):
            if not isinstance(raw_task, dict):
                continue
            title = str(raw_task.get("title") or raw_task.get("name") or "").strip()
            if not title:
                continue
            raw_key = str(raw_task.get("key") or raw_task.get("id") or title)
            key = sanitize_task_key(raw_key, index, used_keys)
            prepared_rows.append((raw_task, key, title))
            alias_map[raw_key.strip().lower()] = key
            alias_map[title.strip().lower()] = key
            alias_map[key] = key

        tasks: list[dict] = []
        seen_keys: set[str] = set()
        for raw_task, key, title in prepared_rows:
            description_text = str(raw_task.get("description") or raw_task.get("summary") or "").strip()
            dependency_keys: list[str] = []
            for dependency in ensure_list(
                raw_task.get("dependency_keys") or raw_task.get("depends_on") or raw_task.get("dependencies")
            ):
                token = str(dependency).strip().lower()
                target_key = alias_map.get(token)
                if target_key and target_key != key and target_key in seen_keys and target_key not in dependency_keys:
                    dependency_keys.append(target_key)

            priority = max(1, min(5, parse_int(raw_task.get("priority"), 2)))
            complexity = max(1, min(5, parse_int(raw_task.get("complexity"), 3)))
            estimate_hours = round(max(0.0, parse_float(raw_task.get("estimate_hours"), 0.0)), 1)
            confidence = round(clamp(parse_float(raw_task.get("confidence"), 0.68), 0.45, 0.95), 2)
            estimate_basis = str(raw_task.get("estimate_basis") or raw_task.get("reason") or "").strip()
            if estimate_hours <= 0:
                estimate_hours, confidence, fallback_basis = estimate_task_hours(
                    title,
                    description_text,
                    complexity,
                    len(dependency_keys),
                    category,
                )
                estimate_basis = estimate_basis or fallback_basis
            if not estimate_basis:
                estimate_basis = f"SiliconFlow {self.model}"

            tasks.append(
                {
                    "key": key,
                    "title": title,
                    "description": description_text,
                    "status": "planned",
                    "owner": "",
                    "priority": priority,
                    "complexity": complexity,
                    "estimate_hours": estimate_hours,
                    "actual_hours": 0.0,
                    "progress": 0,
                    "start_date": start_date,
                    "end_date": start_date,
                    "dependency_keys": dependency_keys,
                    "confidence": confidence,
                    "estimate_basis": estimate_basis,
                    "notes": str(raw_task.get("notes") or "").strip(),
                    "auto_generated": 1,
                }
            )
            seen_keys.add(key)

        if not tasks:
            raise ValueError("SiliconFlow \u4EFB\u52A1\u7ED3\u679C\u4E3A\u7A7A")

        if structured_hint:
            tasks = merge_llm_tasks_with_structured_hint(tasks, structured_hint.get("tasks", []))
        else:
            apply_due_date_pressure(tasks, start_date, due_date)

        project_description = str(parsed.get("project_description") or parsed.get("summary") or "").strip()
        if project_description and len(project_description) > 200:
            project_description = project_description[:200].rstrip("\uFF0C,\uFF1B; ") + "\u2026"

        return {
            "category": category,
            "tasks": tasks,
            "project_description": project_description,
        }

    def analyze_meeting_updates(self, project: dict, tasks: list[dict], meeting_text: str) -> dict:
        compact_tasks = []
        for task in tasks[:120]:
            compact_tasks.append(
                {
                    "task_id": parse_int(task.get("id")),
                    "title": str(task.get("title") or "").strip(),
                    "status": str(task.get("status") or "planned").strip(),
                    "progress": int(clamp(parse_int(task.get("progress"), 0), 0, 100)),
                    "owner": str(task.get("owner") or "").strip(),
                    "description": str(task.get("description") or "").strip()[:180],
                    "notes": str(task.get("notes") or "").strip()[:180],
                }
            )

        input_language = "zh-CN" if contains_cjk(f"{project.get('name', '')} {meeting_text}") else "same-as-user"
        system_prompt = (
            "You are a project meeting assistant. Return JSON only. Do not return Markdown. "
            "Use only task_id values from the provided task list. "
            "Do not create new tasks. Do not guess updates without evidence in meeting notes. "
            "Each update should include task_id, progress(0-100), optional status(planned/in_progress/blocked/done), "
            "short reason, and confidence(0-1)."
        )
        payload = {
            "model": self.model,
            "temperature": 0.1,
            "max_tokens": 1800,
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": system_prompt,
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "task": "Extract progress updates from meeting notes and map them to existing tasks.",
                            "language_preference": input_language,
                            "project": {
                                "name": project.get("name", ""),
                                "description": project.get("description", ""),
                            },
                            "meeting_notes": meeting_text,
                            "tasks": compact_tasks,
                            "requirements": {
                                "rules": [
                                    "Only return updates with clear evidence from meeting notes.",
                                    "task_id must come from provided tasks.",
                                    "progress must be integer 0-100.",
                                    "If status is provided, it must be one of planned/in_progress/blocked/done.",
                                    "Prefer conservative updates when evidence is weak.",
                                ],
                                "output_schema": {
                                    "summary": "One sentence summary",
                                    "updates": [
                                        {
                                            "task_id": 12,
                                            "progress": 65,
                                            "status": "in_progress",
                                            "reason": "完成模块A并开始联调",
                                            "confidence": 0.83,
                                        }
                                    ],
                                },
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        }

        request = Request(
            self.base_url,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.api_key}",
            },
            method="POST",
        )
        with urlopen(request, timeout=self.timeout_seconds) as response:
            response_payload = json.loads(response.read().decode("utf-8"))

        content = ((((response_payload.get("choices") or [{}])[0]).get("message") or {}).get("content") or "").strip()
        if not content:
            raise ValueError("SiliconFlow 返回空响应")

        parsed = extract_json_object(content)
        raw_updates = parsed.get("updates") or parsed.get("task_updates") or parsed.get("items") or []
        if not isinstance(raw_updates, list):
            raise ValueError("SiliconFlow 未返回有效更新列表")

        updates: list[dict] = []
        for raw in raw_updates:
            if not isinstance(raw, dict):
                continue
            task_id = parse_int(raw.get("task_id"), 0)
            if task_id <= 0:
                continue
            if raw.get("progress") in (None, ""):
                continue
            progress = int(clamp(parse_int(raw.get("progress"), 0), 0, 100))
            status = str(raw.get("status") or "").strip()
            if status not in {"planned", "in_progress", "blocked", "done"}:
                status = ""
            reason = str(raw.get("reason") or raw.get("evidence") or "").strip()
            if len(reason) > 180:
                reason = reason[:180].rstrip("，,；; ") + "…"
            confidence = round(clamp(parse_float(raw.get("confidence"), 0.68), 0.0, 1.0), 2)
            updates.append(
                {
                    "task_id": task_id,
                    "progress": progress,
                    "status": status,
                    "reason": reason,
                    "confidence": confidence,
                }
            )

        summary = str(parsed.get("summary") or parsed.get("note") or "").strip()
        if len(summary) > 220:
            summary = summary[:220].rstrip("，,；; ") + "…"

        return {
            "summary": summary,
            "updates": updates,
        }

def build_rule_task_suggestions(
    name: str,
    description: str,
    start_date: str,
    due_date: str | None = None,
) -> tuple[str, list[dict]]:
    category = infer_category(name, description)
    template = PROJECT_TEMPLATES.get(category, PROJECT_TEMPLATES["general"])
    tasks = []
    for item in template:
        dependency_count = len(item["depends_on"])
        estimate_hours, confidence, basis = estimate_task_hours(
            item["title"],
            item["description"],
            item["complexity"],
            dependency_count,
            category,
        )
        tasks.append(
            {
                "key": item["key"],
                "title": item["title"],
                "description": item["description"],
                "status": "planned",
                "owner": "",
                "priority": item["priority"],
                "complexity": item["complexity"],
                "estimate_hours": estimate_hours,
                "actual_hours": 0.0,
                "progress": 0,
                "start_date": start_date,
                "end_date": start_date,
                "dependency_keys": list(item["depends_on"]),
                "confidence": confidence,
                "estimate_basis": basis,
                "notes": "",
                "auto_generated": 1,
            }
        )

    text = f"{name} {description}".lower()
    if any(token in text for token in ("部署", "上线", "docker", "server", "发布")) and not any(
        task["key"] == "ops" for task in tasks
    ):
        estimate_hours, confidence, basis = estimate_task_hours(
            "部署与运行环境",
            "准备运行环境、备份方案和发布步骤。",
            3,
            1,
            category,
        )
        tasks.append(
            {
                "key": "ops",
                "title": "部署与运行环境",
                "description": "准备运行环境、备份方案和发布步骤。",
                "status": "planned",
                "owner": "",
                "priority": 2,
                "complexity": 3,
                "estimate_hours": estimate_hours,
                "actual_hours": 0.0,
                "progress": 0,
                "start_date": start_date,
                "end_date": start_date,
                "dependency_keys": [tasks[-1]["key"]],
                "confidence": confidence,
                "estimate_basis": basis,
                "notes": "",
                "auto_generated": 1,
            }
        )

    apply_due_date_pressure(tasks, start_date, due_date)
    return category, tasks


def build_task_suggestions(
    name: str,
    description: str,
    start_date: str,
    due_date: str | None = None,
    planner: SiliconFlowPlanner | None = None,
) -> dict:
    structured_result = parse_mermaid_gantt_tasks(name, description, start_date, due_date)
    if structured_result is None:
        structured_result = parse_numbered_outline_tasks(name, description, start_date, due_date)
    fallback_category, fallback_tasks = build_rule_task_suggestions(name, description, start_date, due_date)
    result = structured_result or {
        "category": fallback_category,
        "tasks": fallback_tasks,
        "source": "rules",
        "model": "",
        "note": "\u5DF2\u6309\u672C\u5730\u89C4\u5219\u751F\u6210\u4EFB\u52A1\u5EFA\u8BAE",
    }
    if planner is None:
        if structured_result:
            result["note"] = "\u672A\u914D\u7F6E SiliconFlow API Key\uFF0C\u5DF2\u4F7F\u7528\u7ED3\u6784\u5316\u89E3\u6790\u7ED3\u679C"
        else:
            result["note"] = "\u672A\u914D\u7F6E SiliconFlow API Key\uFF0C\u5DF2\u56DE\u9000\u5230\u672C\u5730\u89C4\u5219\u62C6\u5206"
        return result
    try:
        llm_result = planner.build_task_suggestions(
            name,
            description,
            start_date,
            due_date,
            structured_result,
        )
        merged_tasks = llm_result["tasks"]
        payload = {
            "category": structured_result.get("category", llm_result["category"]) if structured_result else llm_result["category"],
            "tasks": merged_tasks,
            "source": "llm",
            "model": planner.model,
            "note": f"已使用 SiliconFlow - {planner.model} 完成项目解析",
        }
        project_description = llm_result.get("project_description", "")
        if structured_result:
            merged_tasks = merge_llm_tasks_with_structured_hint(llm_result["tasks"], structured_result.get("tasks", []))
            payload["tasks"] = merged_tasks
            payload["note"] = f"已使用 SiliconFlow - {planner.model} 完成补充分析，并保留结构化排期"
            payload["project_description"] = structured_result.get("project_description") or project_description
            payload["start_date"] = structured_result.get("start_date")
            payload["due_date"] = structured_result.get("due_date")
            payload["preserve_schedule"] = structured_result.get("preserve_schedule", False)
        elif project_description:
            payload["project_description"] = project_description
        return payload
    except Exception as exc:
        print(f"SiliconFlow fallback: {exc}")
        if structured_result:
            structured_result["note"] = f"SiliconFlow \u8C03\u7528\u5931\u8D25\uFF0C\u5DF2\u4FDD\u7559\u7ED3\u6784\u5316\u89E3\u6790\u7ED3\u679C: {exc}"
            return structured_result
        result["note"] = f"SiliconFlow \u8C03\u7528\u5931\u8D25\uFF0C\u5DF2\u56DE\u9000\u5230\u672C\u5730\u89C4\u5219\u62C6\u5206: {exc}"
        return result


def task_duration_days(estimate_hours: float) -> int:
    return max(1, math.ceil(parse_float(estimate_hours, 6.0) / HOURS_PER_DAY))


def normalize_progress(status: str, current: int | None = None) -> int:
    if current not in (None, ""):
        return int(clamp(parse_int(current, 0), 0, 100))
    mapping = {"planned": 0, "in_progress": 45, "blocked": 20, "done": 100}
    return mapping.get(status, 0)



def infer_status_from_progress(progress: int, current_status: str = "planned") -> str:
    value = int(clamp(parse_int(progress, 0), 0, 100))
    if value >= 100:
        return "done"
    if value <= 0:
        return "planned"
    if current_status == "blocked":
        return "blocked"
    return "in_progress"
def normalize_dependencies(raw_value) -> list[int]:
    if raw_value is None:
        return []
    if isinstance(raw_value, list):
        return [parse_int(item) for item in raw_value if parse_int(item) > 0]
    text = str(raw_value).replace("，", ",").replace("、", ",")
    result = []
    for part in text.split(","):
        value = parse_int(part.strip())
        if value > 0:
            result.append(value)
    return result


def safe_filename(text: str) -> str:
    ascii_only = "".join(char for char in text if char.isascii() and (char.isalnum() or char in ("-", "_")))
    return ascii_only[:48] or "project"


def build_excel_project_sheet(sheet, project: dict) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 42

    header_fill = PatternFill("solid", fgColor="3E474E")
    panel_fill = PatternFill("solid", fgColor="F8F4EC")
    border = Border(
        left=Side(style="thin", color="D7D2C8"),
        right=Side(style="thin", color="D7D2C8"),
        top=Side(style="thin", color="D7D2C8"),
        bottom=Side(style="thin", color="D7D2C8"),
    )

    fields = [
        ("项目 ID", project.get("id")),
        ("项目名称", project.get("name")),
        ("项目描述", project.get("description")),
        ("项目分类", project.get("category")),
        ("项目状态", project.get("status")),
        ("开始日期", parse_date(project.get("start_date"))),
        ("截止日期", parse_date(project.get("due_date"))),
    ]

    sheet.append(["字段", "值"])
    for label, value in fields:
        sheet.append([label, value])

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2, max_row=len(fields) + 1, min_col=1, max_col=2):
        for cell in row:
            cell.fill = panel_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if cell.column == 2 and isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"


def build_excel_task_sheet(sheet, tasks: list[dict]) -> None:
    sheet.freeze_panes = "A2"
    headers = [
        "id",
        "title",
        "description",
        "status",
        "owner",
        "priority",
        "complexity",
        "estimate_hours",
        "actual_hours",
        "progress",
        "start_date",
        "end_date",
        "dependency_ids",
        "dependency_titles",
        "notes",
    ]
    sheet.append(headers)
    for task in tasks:
        sheet.append(
            [
                task["id"],
                task["title"],
                task["description"],
                task["status"],
                task["owner"],
                task["priority"],
                task["complexity"],
                task["estimate_hours"],
                task["actual_hours"],
                task["progress"],
                parse_date(task["start_date"]),
                parse_date(task["end_date"]),
                ",".join(str(item) for item in task["dependency_ids"]),
                ",".join(task["dependency_titles"]),
                task["notes"],
            ]
        )

    widths = {
        "A": 8,
        "B": 28,
        "C": 42,
        "D": 14,
        "E": 14,
        "F": 10,
        "G": 10,
        "H": 12,
        "I": 12,
        "J": 10,
        "K": 13,
        "L": 13,
        "M": 18,
        "N": 26,
        "O": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    header_fill = PatternFill("solid", fgColor="3E474E")
    body_fill = PatternFill("solid", fgColor="FBF8F2")
    border = Border(
        left=Side(style="thin", color="DDD8CF"),
        right=Side(style="thin", color="DDD8CF"),
        top=Side(style="thin", color="DDD8CF"),
        bottom=Side(style="thin", color="DDD8CF"),
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = body_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[9].number_format = '0"%"'
        for date_cell in (row[10], row[11]):
            if isinstance(date_cell.value, date):
                date_cell.number_format = "yyyy-mm-dd"

    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:O{sheet.max_row}"


def _prepare_excel_gantt_sheet(sheet, project: dict, tasks: list[dict], title: str, note: str) -> dict:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "E7"
    sheet.sheet_properties.tabColor = "2B4F8B"

    timeline_start_col = 5
    month_row = 4
    date_row = 5
    header_row = 6
    data_start_row = 7

    project_start = parse_date(project.get("start_date")) or date.today()
    project_due = parse_date(project.get("due_date"))
    task_start_dates = [parse_date(task.get("start_date")) for task in tasks if parse_date(task.get("start_date"))]
    task_end_dates = [parse_date(task.get("end_date")) for task in tasks if parse_date(task.get("end_date"))]
    range_start = min([project_start, *task_start_dates]) if task_start_dates else project_start
    range_end_candidates = [*task_end_dates]
    if project_due:
        range_end_candidates.append(project_due)
    range_end = max(range_end_candidates) if range_end_candidates else range_start + timedelta(days=13)
    if range_end < range_start:
        range_end = range_start
    if (range_end - range_start).days < 13:
        range_end = range_start + timedelta(days=13)

    total_days = (range_end - range_start).days + 1
    timeline_end_col = timeline_start_col + total_days - 1
    end_col_letter = get_column_letter(timeline_end_col)
    weekday_labels = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    for column, width in {"A": 28, "B": 13, "C": 13, "D": 10}.items():
        sheet.column_dimensions[column].width = width
    for col_index in range(timeline_start_col, timeline_end_col + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 4.6

    title_fill = PatternFill(fill_type="solid", start_color="FFF4EEE2", end_color="FFF4EEE2")
    month_fill = PatternFill(fill_type="solid", start_color="FFD8E3F2", end_color="FFD8E3F2")
    day_fill = PatternFill(fill_type="solid", start_color="FFF7FAFC", end_color="FFF7FAFC")
    weekend_fill = PatternFill(fill_type="solid", start_color="FFE7EEF6", end_color="FFE7EEF6")
    header_fill = PatternFill(fill_type="solid", start_color="FF46484C", end_color="FF46484C")
    task_fill = PatternFill(fill_type="solid", start_color="FFFBF8F2", end_color="FFFBF8F2")
    task_fill_alt = PatternFill(fill_type="solid", start_color="FFF3EFE7", end_color="FFF3EFE7")
    project_fill = PatternFill(fill_type="solid", start_color="FFE4D8BF", end_color="FFE4D8BF")
    project_timeline_fill = PatternFill(fill_type="solid", start_color="FFF2E8D6", end_color="FFF2E8D6")
    separator_fill = PatternFill(fill_type="solid", start_color="FFE1D5BF", end_color="FFE1D5BF")
    complete_fill = PatternFill(fill_type="solid", start_color="FF1F437A", end_color="FF1F437A")
    remaining_fill = PatternFill(fill_type="solid", start_color="FF5E7FBC", end_color="FF5E7FBC")
    border = Border(
        left=Side(style="thin", color="FFD3CEC4"),
        right=Side(style="thin", color="FFD3CEC4"),
        top=Side(style="thin", color="FFD3CEC4"),
        bottom=Side(style="thin", color="FFD3CEC4"),
    )
    project_border = Border(
        left=Side(style="thin", color="FFB79B6A"),
        right=Side(style="thin", color="FFB79B6A"),
        top=Side(style="medium", color="FF9F7F4E"),
        bottom=Side(style="medium", color="FF9F7F4E"),
    )
    timeline_border = Border(
        top=Side(style="thin", color="FFD3CEC4"),
        bottom=Side(style="thin", color="FFD3CEC4"),
    )
    project_timeline_border = Border(
        top=Side(style="medium", color="FF9F7F4E"),
        bottom=Side(style="medium", color="FF9F7F4E"),
    )
    separator_border = Border(
        top=Side(style="medium", color="FFBDA176"),
        bottom=Side(style="medium", color="FFBDA176"),
    )
    complete_border = Border(
        left=Side(style="thin", color="FF1F437A"),
        right=Side(style="thin", color="FF1F437A"),
        top=Side(style="thin", color="FF1F437A"),
        bottom=Side(style="thin", color="FF1F437A"),
    )
    remaining_border = Border(
        left=Side(style="thin", color="FF5E7FBC"),
        right=Side(style="thin", color="FF5E7FBC"),
        top=Side(style="thin", color="FF5E7FBC"),
        bottom=Side(style="thin", color="FF5E7FBC"),
    )

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=timeline_end_col)
    sheet.cell(1, 1).value = title
    sheet.cell(1, 1).font = Font(size=18, bold=True, color="FF11232C")
    sheet.cell(1, 1).fill = title_fill
    sheet.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=timeline_end_col)
    sheet.cell(2, 1).value = note
    sheet.cell(2, 1).font = Font(size=10, color="FF556369")
    sheet.cell(2, 1).fill = title_fill
    sheet.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=timeline_end_col)
    sheet.cell(3, 1).value = f"项目周期：{format_date(range_start)} -> {format_date(range_end)}"
    sheet.cell(3, 1).font = Font(size=10, color="FF66757B")
    sheet.cell(3, 1).fill = title_fill
    sheet.cell(3, 1).alignment = Alignment(horizontal="left", vertical="center")

    current_month_start = timeline_start_col
    current_month = range_start.month
    current_year = range_start.year
    for offset in range(total_days):
        current_day = range_start + timedelta(days=offset)
        col_index = timeline_start_col + offset
        is_weekend = current_day.weekday() >= 5
        base_fill = weekend_fill if is_weekend else task_fill

        date_cell = sheet.cell(date_row, col_index)
        date_cell.value = current_day
        date_cell.number_format = "d"
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.font = Font(size=10, bold=True, color="FF22323A")
        date_cell.fill = day_fill
        date_cell.border = border

        weekday_cell = sheet.cell(header_row, col_index)
        weekday_cell.value = weekday_labels[current_day.weekday()]
        weekday_cell.alignment = Alignment(horizontal="center", vertical="center")
        weekday_cell.font = Font(size=10, bold=True, color="FFFFFFFF")
        weekday_cell.fill = header_fill
        weekday_cell.border = border

        if offset == total_days - 1 or (current_day + timedelta(days=1)).month != current_month:
            sheet.merge_cells(
                start_row=month_row,
                start_column=current_month_start,
                end_row=month_row,
                end_column=col_index,
            )
            month_cell = sheet.cell(month_row, current_month_start)
            month_cell.value = f"{current_year}-{current_month:02d}"
            month_cell.alignment = Alignment(horizontal="center", vertical="center")
            month_cell.font = Font(size=11, bold=True, color="FF294559")
            month_cell.fill = month_fill
            month_cell.border = border
            current_month_start = col_index + 1
            next_day = current_day + timedelta(days=1)
            current_month = next_day.month
            current_year = next_day.year

        for row_index in range(data_start_row, data_start_row + max(len(tasks), 1)):
            timeline_cell = sheet.cell(row_index, col_index)
            timeline_cell.fill = base_fill
            timeline_cell.border = timeline_border
            timeline_cell.alignment = Alignment(horizontal="center", vertical="center")
            timeline_cell.number_format = ";;;"

    left_headers = ["任务名称", "开始时间", "结束时间", "完成进度"]
    for index, label in enumerate(left_headers, start=1):
        header_cell = sheet.cell(header_row, index)
        header_cell.value = label
        header_cell.font = Font(size=10, bold=True, color="FFFFFFFF")
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = border

    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[3].height = 20
    sheet.row_dimensions[4].height = 21
    sheet.row_dimensions[5].height = 20
    sheet.row_dimensions[6].height = 22

    return {
        "timeline_start_col": timeline_start_col,
        "timeline_end_col": timeline_end_col,
        "end_col_letter": end_col_letter,
        "data_start_row": data_start_row,
        "range_start": range_start,
        "range_end": range_end,
        "total_days": total_days,
        "task_fill": task_fill,
        "task_fill_alt": task_fill_alt,
        "project_fill": project_fill,
        "project_timeline_fill": project_timeline_fill,
        "separator_fill": separator_fill,
        "complete_fill": complete_fill,
        "remaining_fill": remaining_fill,
        "border": border,
        "project_border": project_border,
        "separator_border": separator_border,
        "complete_border": complete_border,
        "remaining_border": remaining_border,
        "timeline_border": timeline_border,
        "project_timeline_border": project_timeline_border,
    }


def _write_excel_gantt_task_row(sheet, frame: dict, row_index: int, row_offset: int, task: dict, is_parent: bool) -> tuple[date | None, date | None, int]:
    row_type = task.get("row_type") or "task"
    row_fill = frame["task_fill"] if row_offset % 2 == 0 else frame["task_fill_alt"]
    row_border = frame["border"]
    title_font = Font(bold=is_parent)
    row_height = 24

    if row_type == "project_header":
        row_fill = frame["project_fill"]
        row_border = frame["project_border"]
        title_font = Font(bold=True, size=11, color="FF1C2D35")
        row_height = 28
    elif row_type == "separator":
        row_fill = frame["separator_fill"]
        row_border = frame["separator_border"]
        title_font = Font(color="FF7D6A48")
        row_height = 12

    start_value = parse_date(task.get("start_date"))
    end_value = parse_date(task.get("end_date"))
    progress_value = int(clamp(parse_int(task.get("progress"), 0), 0, 100)) if row_type != "separator" else 0

    title_cell = sheet.cell(row_index, 1)
    title_cell.value = task.get("title") or ""
    title_cell.alignment = Alignment(vertical="center")
    title_cell.font = title_font
    title_cell.fill = row_fill
    title_cell.border = row_border

    start_cell = sheet.cell(row_index, 2)
    start_cell.value = start_value if row_type != "separator" else None
    start_cell.number_format = "yyyy-mm-dd"
    start_cell.alignment = Alignment(horizontal="center", vertical="center")
    start_cell.fill = row_fill
    start_cell.border = row_border

    end_cell = sheet.cell(row_index, 3)
    end_cell.value = end_value if row_type != "separator" else None
    end_cell.number_format = "yyyy-mm-dd"
    end_cell.alignment = Alignment(horizontal="center", vertical="center")
    end_cell.fill = row_fill
    end_cell.border = row_border

    progress_cell = sheet.cell(row_index, 4)
    progress_cell.value = progress_value if row_type != "separator" else None
    progress_cell.number_format = '0"%"'
    progress_cell.alignment = Alignment(horizontal="center", vertical="center")
    progress_cell.fill = row_fill
    progress_cell.border = row_border

    sheet.row_dimensions[row_index].height = row_height
    return start_value, end_value, progress_value


def _progress_segment_info(start_value: date | None, end_value: date | None, progress_value: int) -> tuple[int, int]:
    if not start_value or not end_value or end_value < start_value:
        return 0, 0
    total_duration = (end_value - start_value).days + 1
    if progress_value <= 0:
        return total_duration, 0
    complete_days = min(total_duration, math.ceil(total_duration * progress_value / 100))
    return total_duration, complete_days


def _merge_excel_bar_segment(sheet, row_index: int, start_col: int, end_col: int, fill, border) -> None:
    if start_col > end_col:
        return
    if end_col > start_col:
        sheet.merge_cells(start_row=row_index, start_column=start_col, end_row=row_index, end_column=end_col)
    anchor = sheet.cell(row_index, start_col)
    anchor.value = ""
    anchor.fill = fill
    anchor.border = border
    anchor.alignment = Alignment(horizontal="center", vertical="center")


def _build_excel_dynamic_gantt_sheet(
    sheet,
    project: dict,
    tasks: list[dict],
    sheet_title: str = "甘特图",
    title: str | None = None,
    note: str | None = None,
) -> None:
    sheet.title = sheet_title
    frame = _prepare_excel_gantt_sheet(
        sheet,
        project,
        tasks,
        title or f"{project.get('name') or '项目'} 动态甘特图",
        note or (project.get("description") or "任务排期视图"),
    )

    if not tasks:
        empty_cells = [sheet.cell(frame["data_start_row"], column) for column in range(1, 5)]
        empty_cells[0].value = "暂无任务"
        for cell in empty_cells:
            cell.fill = frame["task_fill"]
            cell.border = frame["border"]
            cell.alignment = Alignment(vertical="center")
        return

    parent_ids = {task.get("parent_id") for task in tasks if task.get("parent_id")}
    for offset, task in enumerate(tasks):
        row_index = frame["data_start_row"] + offset
        row_type = task.get("row_type") or "task"
        start_value, end_value, progress_value = _write_excel_gantt_task_row(
            sheet,
            frame,
            row_index,
            offset,
            task,
            task["id"] in parent_ids,
        )
        total_duration, complete_days = _progress_segment_info(start_value, end_value, progress_value)

        for day_offset in range(frame["total_days"]):
            current_day = frame["range_start"] + timedelta(days=day_offset)
            col_index = frame["timeline_start_col"] + day_offset
            col_letter = get_column_letter(col_index)
            timeline_cell = sheet.cell(row_index, col_index)
            timeline_cell.number_format = ";;;"

            if row_type == "separator":
                timeline_cell.value = ""
                timeline_cell.fill = frame["separator_fill"]
                timeline_cell.border = frame["separator_border"]
                continue

            if row_type == "project_header":
                timeline_cell.fill = frame["project_timeline_fill"]
                timeline_cell.border = frame["project_timeline_border"]
            else:
                timeline_cell.border = frame["timeline_border"]

            timeline_cell.value = (
                f'=IF(OR(NOT(ISNUMBER($B{row_index})),NOT(ISNUMBER($C{row_index})),{col_letter}$5<$B{row_index},{col_letter}$5>$C{row_index}),"",'
                f'IF(AND($D{row_index}>0,{col_letter}$5<=$B{row_index}+ROUNDUP((($C{row_index}-$B{row_index}+1)*$D{row_index}/100),0)-1),2,1))'
            )
            if total_duration and start_value and end_value and start_value <= current_day <= end_value:
                complete_end = start_value + timedelta(days=complete_days - 1) if complete_days else None
                timeline_cell.fill = frame["complete_fill"] if complete_end and current_day <= complete_end else frame["remaining_fill"]

    data_end_row = frame["data_start_row"] + len(tasks) - 1
    progress_range = f"D{frame['data_start_row']}:D{data_end_row}"
    timeline_range = f"E{frame['data_start_row']}:{frame['end_col_letter']}{data_end_row}"
    sheet.conditional_formatting.add(
        progress_range,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="FF1F437A"),
    )
    sheet.conditional_formatting.add(
        timeline_range,
        CellIsRule(operator="equal", formula=["2"], fill=frame["complete_fill"], stopIfTrue=True),
    )
    sheet.conditional_formatting.add(
        timeline_range,
        CellIsRule(operator="equal", formula=["1"], fill=frame["remaining_fill"]),
    )
    sheet.auto_filter.ref = f"A6:D{data_end_row}"


def _build_excel_snapshot_gantt_sheet(sheet, project: dict, tasks: list[dict]) -> None:
    sheet.title = "甘特图"
    frame = _prepare_excel_gantt_sheet(
        sheet,
        project,
        tasks,
        f"{project.get('name') or '项目'} 甘特图展示版",
        "展示版：连续时间条已合并，适合查看和汇报。如需在 Excel 内联动编辑，请使用“甘特图-动态”工作表。",
    )

    if not tasks:
        empty_cells = [sheet.cell(frame["data_start_row"], column) for column in range(1, 5)]
        empty_cells[0].value = "暂无任务"
        for cell in empty_cells:
            cell.fill = frame["task_fill"]
            cell.border = frame["border"]
            cell.alignment = Alignment(vertical="center")
        return

    parent_ids = {task.get("parent_id") for task in tasks if task.get("parent_id")}
    for offset, task in enumerate(tasks):
        row_index = frame["data_start_row"] + offset
        start_value, end_value, progress_value = _write_excel_gantt_task_row(
            sheet,
            frame,
            row_index,
            offset,
            task,
            task["id"] in parent_ids,
        )
        total_duration, complete_days = _progress_segment_info(start_value, end_value, progress_value)
        if not total_duration or not start_value or not end_value:
            continue

        start_col = frame["timeline_start_col"] + (start_value - frame["range_start"]).days
        end_col = frame["timeline_start_col"] + (end_value - frame["range_start"]).days

        if complete_days > 0:
            complete_end_col = start_col + complete_days - 1
            _merge_excel_bar_segment(
                sheet,
                row_index,
                start_col,
                complete_end_col,
                frame["complete_fill"],
                frame["complete_border"],
            )
        if complete_days < total_duration:
            remaining_start_col = start_col + complete_days
            _merge_excel_bar_segment(
                sheet,
                row_index,
                remaining_start_col,
                end_col,
                frame["remaining_fill"],
                frame["remaining_border"],
            )

    data_end_row = frame["data_start_row"] + len(tasks) - 1
    progress_range = f"D{frame['data_start_row']}:D{data_end_row}"
    sheet.conditional_formatting.add(
        progress_range,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="FF1F437A"),
    )
    sheet.auto_filter.ref = f"A6:D{data_end_row}"


def build_excel_gantt_sheet(workbook, project: dict, tasks: list[dict]) -> None:
    _build_excel_dynamic_gantt_sheet(workbook.active, project, tasks)


def build_excel_all_task_sheet(sheet, details: list[dict]) -> None:
    sheet.freeze_panes = "A2"
    headers = [
        "项目ID", "项目名称", "项目分类", "任务ID", "任务名称", "描述", "状态", "负责人",
        "优先级", "复杂度", "预估工时", "实际工时", "进度", "开始日期", "结束日期", "依赖任务", "备注",
    ]
    sheet.append(headers)
    for detail in details:
        project = detail["project"]
        for task in detail["tasks"]:
            sheet.append(
                [
                    project.get("id"),
                    project.get("name"),
                    project.get("category"),
                    task.get("id"),
                    task.get("title"),
                    task.get("description"),
                    task.get("status"),
                    task.get("owner"),
                    task.get("priority"),
                    task.get("complexity"),
                    task.get("estimate_hours"),
                    task.get("actual_hours"),
                    task.get("progress"),
                    parse_date(task.get("start_date")),
                    parse_date(task.get("end_date")),
                    ",".join(task.get("dependency_titles", [])),
                    task.get("notes"),
                ]
            )

    widths = {"A": 10, "B": 22, "C": 12, "D": 8, "E": 28, "F": 40, "G": 12, "H": 14, "I": 10, "J": 10, "K": 12, "L": 12, "M": 10, "N": 13, "O": 13, "P": 26, "Q": 28}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    header_fill = PatternFill("solid", fgColor="3E474E")
    body_fill = PatternFill("solid", fgColor="FBF8F2")
    border = Border(
        left=Side(style="thin", color="DDD8CF"),
        right=Side(style="thin", color="DDD8CF"),
        top=Side(style="thin", color="DDD8CF"),
        bottom=Side(style="thin", color="DDD8CF"),
    )
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = body_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[12].number_format = '0"%"'
        for date_cell in (row[13], row[14]):
            if isinstance(date_cell.value, date):
                date_cell.number_format = "yyyy-mm-dd"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:Q{sheet.max_row}"


def build_excel_project_list_sheet(sheet, details: list[dict]) -> None:
    sheet.freeze_panes = "A2"
    headers = ["项目ID", "项目名称", "分类", "状态", "开始日期", "截止日期", "分析来源", "分析模型", "任务总数", "已完成", "阻塞", "预计完成", "延期天数"]
    sheet.append(headers)
    for detail in details:
        project = detail["project"]
        stats = detail["stats"]
        sheet.append([
            project.get("id"),
            project.get("name"),
            project.get("category"),
            project.get("status"),
            parse_date(project.get("start_date")),
            parse_date(project.get("due_date")),
            project.get("analysis_source"),
            project.get("analysis_model"),
            stats.get("total_tasks"),
            stats.get("completed_tasks"),
            stats.get("blocked_tasks"),
            parse_date(stats.get("projected_finish")),
            stats.get("slip_days"),
        ])

    widths = {"A": 10, "B": 24, "C": 12, "D": 12, "E": 13, "F": 13, "G": 12, "H": 20, "I": 10, "J": 10, "K": 10, "L": 13, "M": 10}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    header_fill = PatternFill("solid", fgColor="3E474E")
    body_fill = PatternFill("solid", fgColor="FBF8F2")
    border = Border(
        left=Side(style="thin", color="DDD8CF"),
        right=Side(style="thin", color="DDD8CF"),
        top=Side(style="thin", color="DDD8CF"),
        bottom=Side(style="thin", color="DDD8CF"),
    )
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = body_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for date_cell in (row[4], row[5], row[11]):
            if isinstance(date_cell.value, date):
                date_cell.number_format = "yyyy-mm-dd"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:M{sheet.max_row}"


def build_excel_all_projects_gantt_sheet(workbook, details: list[dict]) -> None:
    details = sorted(
        details,
        key=lambda detail: (
            parse_date(detail["project"].get("start_date")) or date.today(),
            parse_date(detail["project"].get("due_date")) or date.today(),
            parse_int(detail["project"].get("id"), 0),
        ),
    )
    combined_tasks: list[dict] = []
    overall_starts: list[date] = []
    overall_ends: list[date] = []
    synthetic_base = 900000

    for offset, detail in enumerate(details, start=1):
        project = detail["project"]
        stats = detail["stats"]
        project_start = project.get("start_date") or today_str()
        project_end = project.get("due_date") or stats.get("projected_finish") or project_start
        start_value = parse_date(project_start)
        end_value = parse_date(project_end)
        if start_value:
            overall_starts.append(start_value)
        if end_value:
            overall_ends.append(end_value)

        total_tasks = parse_int(stats.get("total_tasks"), 0)
        completed_tasks = parse_int(stats.get("completed_tasks"), 0)
        project_progress = int(round(completed_tasks * 100 / total_tasks)) if total_tasks else 0
        parent_id = synthetic_base + offset
        combined_tasks.append(
            {
                "id": parent_id,
                "parent_id": None,
                "row_type": "project_header",
                "title": f"【{project.get('name') or '未命名项目'}】",
                "description": str(project.get("description") or "")[:120],
                "status": project.get("status") or "planned",
                "owner": "",
                "priority": 1,
                "complexity": 1,
                "estimate_hours": round(parse_float(stats.get("estimated_hours"), 0.0), 1),
                "actual_hours": round(parse_float(stats.get("actual_hours"), 0.0), 1),
                "progress": project_progress,
                "start_date": project_start,
                "end_date": project_end,
                "dependency_ids": [],
                "dependency_titles": [],
                "confidence": 1.0,
                "estimate_basis": "项目汇总",
                "notes": f"项目分类：{project.get('category') or 'general'}；任务数：{total_tasks}",
            }
        )
        for task in detail["tasks"]:
            merged_task = dict(task)
            merged_task["row_type"] = "task"
            merged_task["title"] = f"    {task.get('title') or ''}"
            merged_task["parent_id"] = parent_id
            note_parts = [f"所属项目：{project.get('name') or ''}"]
            if task.get("notes"):
                note_parts.append(str(task.get("notes")))
            merged_task["notes"] = " | ".join(part for part in note_parts if part)
            combined_tasks.append(merged_task)
        if offset < len(details):
            combined_tasks.append(
                {
                    "id": synthetic_base + 50000 + offset,
                    "parent_id": None,
                    "row_type": "separator",
                    "title": "",
                    "description": "",
                    "status": "planned",
                    "owner": "",
                    "priority": 0,
                    "complexity": 0,
                    "estimate_hours": 0,
                    "actual_hours": 0,
                    "progress": 0,
                    "start_date": None,
                    "end_date": None,
                    "dependency_ids": [],
                    "dependency_titles": [],
                    "confidence": 0.0,
                    "estimate_basis": "",
                    "notes": "",
                }
            )

    project_start = format_date(min(overall_starts)) if overall_starts else today_str()
    project_due = format_date(max(overall_ends)) if overall_ends else project_start
    pseudo_project = {
        "name": "全部项目",
        "description": "所有项目任务的统一甘特排期视图",
        "start_date": project_start,
        "due_date": project_due,
    }
    _build_excel_dynamic_gantt_sheet(
        workbook.active,
        pseudo_project,
        combined_tasks,
        sheet_title="全部项目甘特图",
        title="全部项目甘特图",
        note="按项目分组汇总全部任务，可直接在表中查看整体排期。",
    )


def row_to_dict(row: sqlite3.Row) -> dict:
    payload = dict(row)
    payload["auto_generated"] = int(payload.get("auto_generated", 0))
    payload["priority"] = int(payload.get("priority", 0))
    payload["complexity"] = int(payload.get("complexity", 0))
    payload["progress"] = int(payload.get("progress", 0))
    payload["estimate_hours"] = round(parse_float(payload.get("estimate_hours")), 1)
    payload["actual_hours"] = round(parse_float(payload.get("actual_hours")), 1)
    payload["confidence"] = round(parse_float(payload.get("confidence"), 0.56), 2)
    return payload


class TaskDatabase:
    def __init__(self, db_path: Path, planner: SiliconFlowPlanner | None = None):
        self.db_path = db_path
        self.planner = planner
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()

    def connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.db_path)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        return connection

    def _init_db(self) -> None:
        with self.connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS projects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    description TEXT DEFAULT '',
                    category TEXT DEFAULT 'general',
                    status TEXT DEFAULT 'planned',
                    start_date TEXT NOT NULL,
                    due_date TEXT,
                    analysis_source TEXT DEFAULT 'manual',
                    analysis_model TEXT DEFAULT '',
                    analysis_note TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS tasks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    parent_id INTEGER,
                    title TEXT NOT NULL,
                    description TEXT DEFAULT '',
                    status TEXT DEFAULT 'planned',
                    owner TEXT DEFAULT '',
                    priority INTEGER DEFAULT 2,
                    complexity INTEGER DEFAULT 3,
                    estimate_hours REAL DEFAULT 6,
                    actual_hours REAL DEFAULT 0,
                    progress INTEGER DEFAULT 0,
                    start_date TEXT,
                    end_date TEXT,
                    confidence REAL DEFAULT 0.56,
                    estimate_basis TEXT DEFAULT '',
                    notes TEXT DEFAULT '',
                    auto_generated INTEGER DEFAULT 0,
                    sort_order INTEGER DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
                    FOREIGN KEY(parent_id) REFERENCES tasks(id) ON DELETE SET NULL
                );

                CREATE TABLE IF NOT EXISTS task_dependencies (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id INTEGER NOT NULL,
                    depends_on_task_id INTEGER NOT NULL,
                    UNIQUE(task_id, depends_on_task_id),
                    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                    FOREIGN KEY(depends_on_task_id) REFERENCES tasks(id) ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS estimate_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id INTEGER NOT NULL,
                    estimate_hours REAL NOT NULL,
                    confidence REAL NOT NULL,
                    basis TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS project_snapshots (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    action TEXT NOT NULL,
                    summary TEXT DEFAULT '',
                    task_count INTEGER DEFAULT 0,
                    payload_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS import_export_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER,
                    action TEXT NOT NULL,
                    file_format TEXT NOT NULL,
                    row_count INTEGER DEFAULT 0,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE SET NULL
                );
                """
            )
            self._ensure_project_columns(connection)
            project_count = connection.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
            if project_count == 0:
                self._seed(connection)
            self._ensure_initial_snapshots(connection)

    def _ensure_project_columns(self, connection: sqlite3.Connection) -> None:
        columns = {row[1] for row in connection.execute("PRAGMA table_info(projects)").fetchall()}
        if "analysis_source" not in columns:
            connection.execute("ALTER TABLE projects ADD COLUMN analysis_source TEXT DEFAULT 'manual'")
        if "analysis_model" not in columns:
            connection.execute("ALTER TABLE projects ADD COLUMN analysis_model TEXT DEFAULT ''")
        if "analysis_note" not in columns:
            connection.execute("ALTER TABLE projects ADD COLUMN analysis_note TEXT DEFAULT ''")
        if "deleted_at" not in columns:
            connection.execute("ALTER TABLE projects ADD COLUMN deleted_at TEXT DEFAULT NULL")

    def _ensure_initial_snapshots(self, connection: sqlite3.Connection) -> None:
        project_rows = connection.execute(
            """
            SELECT p.id
            FROM projects p
            LEFT JOIN project_snapshots ps ON ps.project_id = p.id
            WHERE p.deleted_at IS NULL
            GROUP BY p.id
            HAVING COUNT(ps.id) = 0
            """
        ).fetchall()
        for row in project_rows:
            self._create_project_snapshot(
                connection,
                parse_int(row["id"], 0),
                "baseline",
                "启用历史回退时自动创建的基线快照",
            )
    def _seed(self, connection: sqlite3.Connection) -> None:
        start_date = today_str()
        due_date = add_days(start_date, 20)
        category, suggestions = build_rule_task_suggestions(
            "任务甘特图工具第一版",
            "本地网页管理任务，支持项目识别、估时、导入导出和甘特图排期。",
            start_date,
            due_date,
        )
        project_id = self._insert_project(
            connection,
            {
                "name": "任务甘特图工具第一版",
                "description": "示例项目：展示自动拆分、排期和导入导出。",
                "category": category,
                "status": "in_progress",
                "start_date": start_date,
                "due_date": due_date,
            },
        )
        self._insert_tasks_for_project(connection, project_id, suggestions)
        self._recalculate_project(connection, project_id)
        self._create_project_snapshot(connection, project_id, "seed", "初始化示例项目")

    def _insert_project(self, connection: sqlite3.Connection, payload: dict) -> int:
        timestamp = now_iso()
        cursor = connection.execute(
            """
            INSERT INTO projects (
                name, description, category, status, start_date, due_date,
                analysis_source, analysis_model, analysis_note, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload["name"],
                payload.get("description", ""),
                payload.get("category", "general"),
                payload.get("status", "planned"),
                payload.get("start_date", today_str()),
                payload.get("due_date"),
                payload.get("analysis_source", "manual"),
                payload.get("analysis_model", ""),
                payload.get("analysis_note", ""),
                timestamp,
                timestamp,
            ),
        )
        return int(cursor.lastrowid)

    def _insert_estimate_history(
        self, connection: sqlite3.Connection, task_id: int, estimate_hours: float, confidence: float, basis: str
    ) -> None:
        connection.execute(
            """
            INSERT INTO estimate_history (task_id, estimate_hours, confidence, basis, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (task_id, estimate_hours, confidence, basis, now_iso()),
        )

    def _replace_dependencies(self, connection: sqlite3.Connection, task_id: int, dependency_ids: list[int]) -> None:
        connection.execute("DELETE FROM task_dependencies WHERE task_id = ?", (task_id,))
        for dependency_id in sorted(set(dependency_ids)):
            if dependency_id != task_id:
                connection.execute(
                    "INSERT OR IGNORE INTO task_dependencies (task_id, depends_on_task_id) VALUES (?, ?)",
                    (task_id, dependency_id),
                )

    def _serialize_project_snapshot(self, connection: sqlite3.Connection, project_id: int) -> dict:
        detail = self._load_project_detail(connection, project_id)
        project = detail["project"]
        payload_project = {
            "name": project.get("name"),
            "description": project.get("description"),
            "category": project.get("category"),
            "status": project.get("status"),
            "start_date": project.get("start_date"),
            "due_date": project.get("due_date"),
            "analysis_source": project.get("analysis_source"),
            "analysis_model": project.get("analysis_model"),
            "analysis_note": project.get("analysis_note"),
            "created_at": project.get("created_at"),
        }
        payload_tasks: list[dict] = []
        for task in detail["tasks"]:
            payload_tasks.append(
                {
                    "id": parse_int(task.get("id"), 0),
                    "parent_id": parse_int(task.get("parent_id"), 0) or None,
                    "title": task.get("title", ""),
                    "description": task.get("description", ""),
                    "status": task.get("status", "planned"),
                    "owner": task.get("owner", ""),
                    "priority": parse_int(task.get("priority"), 2),
                    "complexity": parse_int(task.get("complexity"), 3),
                    "estimate_hours": parse_float(task.get("estimate_hours"), 0.0),
                    "actual_hours": parse_float(task.get("actual_hours"), 0.0),
                    "progress": parse_int(task.get("progress"), 0),
                    "start_date": task.get("start_date"),
                    "end_date": task.get("end_date"),
                    "confidence": parse_float(task.get("confidence"), 0.56),
                    "estimate_basis": task.get("estimate_basis", ""),
                    "notes": task.get("notes", ""),
                    "auto_generated": parse_int(task.get("auto_generated"), 0),
                    "sort_order": parse_int(task.get("sort_order"), 0),
                    "created_at": task.get("created_at"),
                    "updated_at": task.get("updated_at"),
                    "dependency_ids": [parse_int(item) for item in task.get("dependency_ids", []) if parse_int(item) > 0],
                }
            )
        return {"project": payload_project, "tasks": payload_tasks}

    def _list_project_snapshots_connection(
        self, connection: sqlite3.Connection, project_id: int, limit: int = 30
    ) -> list[dict]:
        rows = connection.execute(
            """
            SELECT id, project_id, action, summary, task_count, created_at
            FROM project_snapshots
            WHERE project_id = ?
            ORDER BY id DESC
            LIMIT ?
            """,
            (project_id, max(1, min(parse_int(limit, 30), 100))),
        ).fetchall()
        snapshots = []
        for row in rows:
            item = dict(row)
            item["task_count"] = parse_int(item.get("task_count"), 0)
            snapshots.append(item)
        return snapshots

    def _create_project_snapshot(
        self, connection: sqlite3.Connection, project_id: int, action: str, summary: str = ""
    ) -> int:
        payload = self._serialize_project_snapshot(connection, project_id)
        cursor = connection.execute(
            """
            INSERT INTO project_snapshots (project_id, action, summary, task_count, payload_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                action,
                summary,
                len(payload.get("tasks", [])),
                json.dumps(payload, ensure_ascii=False),
                now_iso(),
            ),
        )
        snapshot_id = int(cursor.lastrowid)
        stale_rows = connection.execute(
            "SELECT id FROM project_snapshots WHERE project_id = ? ORDER BY id DESC",
            (project_id,),
        ).fetchall()
        stale_ids = [int(row["id"]) for row in stale_rows[80:]]
        if stale_ids:
            placeholders = ",".join("?" for _ in stale_ids)
            connection.execute(
                f"DELETE FROM project_snapshots WHERE id IN ({placeholders})",
                stale_ids,
            )
        return snapshot_id

    def _restore_snapshot_tasks(
        self, connection: sqlite3.Connection, project_id: int, snapshot_tasks: list[dict]
    ) -> None:
        connection.execute("DELETE FROM tasks WHERE project_id = ?", (project_id,))
        task_ids: set[int] = set()
        for task in snapshot_tasks:
            task_id = parse_int(task.get("id"), 0)
            if task_id <= 0:
                continue
            task_ids.add(task_id)
            connection.execute(
                """
                INSERT INTO tasks (
                    id, project_id, parent_id, title, description, status, owner, priority, complexity,
                    estimate_hours, actual_hours, progress, start_date, end_date, confidence,
                    estimate_basis, notes, auto_generated, sort_order, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    task_id,
                    project_id,
                    parse_int(task.get("parent_id"), 0) or None,
                    task.get("title", ""),
                    task.get("description", ""),
                    task.get("status", "planned"),
                    task.get("owner", ""),
                    parse_int(task.get("priority"), 2),
                    parse_int(task.get("complexity"), 3),
                    parse_float(task.get("estimate_hours"), 0.0),
                    parse_float(task.get("actual_hours"), 0.0),
                    normalize_progress(task.get("status", "planned"), task.get("progress")),
                    task.get("start_date"),
                    task.get("end_date"),
                    parse_float(task.get("confidence"), 0.56),
                    task.get("estimate_basis", ""),
                    task.get("notes", ""),
                    parse_int(task.get("auto_generated"), 0),
                    parse_int(task.get("sort_order"), 0),
                    task.get("created_at") or now_iso(),
                    now_iso(),
                ),
            )
            self._insert_estimate_history(
                connection,
                task_id,
                parse_float(task.get("estimate_hours"), 0.0),
                parse_float(task.get("confidence"), 0.56),
                "恢复快照",
            )

        for task in snapshot_tasks:
            task_id = parse_int(task.get("id"), 0)
            if task_id <= 0 or task_id not in task_ids:
                continue
            dependency_ids = [
                dependency_id
                for dependency_id in [parse_int(item) for item in task.get("dependency_ids", [])]
                if dependency_id > 0 and dependency_id in task_ids and dependency_id != task_id
            ]
            self._replace_dependencies(connection, task_id, dependency_ids)

    def list_project_snapshots(self, project_id: int, limit: int = 30) -> list[dict]:
        with self.connect() as connection:
            row = connection.execute(
                "SELECT id FROM projects WHERE id = ? AND deleted_at IS NULL",
                (project_id,),
            ).fetchone()
            if not row:
                raise KeyError("project not found")
            return self._list_project_snapshots_connection(connection, project_id, limit)

    def restore_project_snapshot(self, project_id: int, snapshot_id: int) -> dict:
        with self.connect() as connection:
            project_row = connection.execute(
                "SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL",
                (project_id,),
            ).fetchone()
            if not project_row:
                raise KeyError("project not found")

            snapshot_row = connection.execute(
                "SELECT * FROM project_snapshots WHERE id = ? AND project_id = ?",
                (snapshot_id, project_id),
            ).fetchone()
            if not snapshot_row:
                raise KeyError("snapshot not found")

            self._create_project_snapshot(
                connection,
                project_id,
                "before_restore",
                f"恢复快照 #{snapshot_id} 前自动保存",
            )

            payload = json.loads(snapshot_row["payload_json"])
            snapshot_project = payload.get("project") or {}
            snapshot_tasks = payload.get("tasks") or []

            connection.execute(
                """
                UPDATE projects
                SET name = ?, description = ?, category = ?, status = ?, start_date = ?, due_date = ?,
                    analysis_source = ?, analysis_model = ?, analysis_note = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    snapshot_project.get("name") or project_row["name"],
                    snapshot_project.get("description", ""),
                    snapshot_project.get("category") or project_row["category"],
                    snapshot_project.get("status") or project_row["status"],
                    snapshot_project.get("start_date") or project_row["start_date"],
                    snapshot_project.get("due_date"),
                    snapshot_project.get("analysis_source", "manual"),
                    snapshot_project.get("analysis_model", ""),
                    snapshot_project.get("analysis_note", ""),
                    now_iso(),
                    project_id,
                ),
            )

            self._restore_snapshot_tasks(connection, project_id, snapshot_tasks)
            connection.execute(
                """
                INSERT INTO import_export_log (project_id, action, file_format, row_count, created_at)
                VALUES (?, 'restore_snapshot', ?, ?, ?)
                """,
                (project_id, 'snapshot', len(snapshot_tasks), now_iso()),
            )
            restore_summary = f"已恢复到快照 #{snapshot_id}"
            self._create_project_snapshot(connection, project_id, "restore_snapshot", restore_summary)
            detail = self._load_project_detail(connection, project_id)
            detail["snapshot_restore"] = {
                "snapshot_id": snapshot_id,
                "created_at": snapshot_row["created_at"],
                "summary": snapshot_row["summary"],
                "task_count": parse_int(snapshot_row["task_count"], len(snapshot_tasks)),
            }
            return detail

    def _insert_tasks_for_project(
        self, connection: sqlite3.Connection, project_id: int, tasks: list[dict], sort_offset: int = 0
    ) -> list[int]:
        key_map: dict[str, int] = {}
        parent_ref_map: dict[int, str | int | None] = {}
        created_ids: list[int] = []
        for index, task in enumerate(tasks):
            timestamp = now_iso()
            estimate_hours = parse_float(task.get("estimate_hours"), 6.0)
            cursor = connection.execute(
                """
                INSERT INTO tasks (
                    project_id, parent_id, title, description, status, owner, priority, complexity,
                    estimate_hours, actual_hours, progress, start_date, end_date, confidence,
                    estimate_basis, notes, auto_generated, sort_order, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    None,
                    task["title"],
                    task.get("description", ""),
                    task.get("status", "planned"),
                    task.get("owner", ""),
                    parse_int(task.get("priority"), 2),
                    parse_int(task.get("complexity"), 3),
                    estimate_hours,
                    parse_float(task.get("actual_hours"), 0.0),
                    normalize_progress(task.get("status", "planned"), task.get("progress")),
                    task.get("start_date"),
                    task.get("end_date"),
                    parse_float(task.get("confidence"), 0.56),
                    task.get("estimate_basis", ""),
                    task.get("notes", ""),
                    parse_int(task.get("auto_generated"), 0),
                    sort_offset + index,
                    timestamp,
                    timestamp,
                ),
            )
            task_id = int(cursor.lastrowid)
            created_ids.append(task_id)
            if task.get("key"):
                key_map[str(task["key"])] = task_id
            if task.get("parent_ref") not in (None, ""):
                parent_ref_map[task_id] = task.get("parent_ref")
            self._insert_estimate_history(
                connection,
                task_id,
                estimate_hours,
                parse_float(task.get("confidence"), 0.56),
                task.get("estimate_basis", "创建任务"),
            )

        for index, task in enumerate(tasks):
            task_id = created_ids[index]
            parent_ref = parent_ref_map.get(task_id)
            if parent_ref not in (None, ""):
                parent_id = key_map.get(str(parent_ref), parse_int(parent_ref) or None)
                if parent_id and parent_id != task_id:
                    connection.execute("UPDATE tasks SET parent_id = ? WHERE id = ?", (parent_id, task_id))

            dependency_ids = []
            for dependency in task.get("dependency_keys", []):
                target_id = key_map.get(str(dependency), parse_int(dependency))
                if target_id:
                    dependency_ids.append(target_id)
            self._replace_dependencies(connection, task_id, dependency_ids)
        return created_ids

    def list_projects(self, deleted: bool = False) -> list[dict]:
        where_clause = "p.deleted_at IS NOT NULL" if deleted else "p.deleted_at IS NULL"
        order_column = "p.deleted_at" if deleted else "p.updated_at"
        with self.connect() as connection:
            rows = connection.execute(
                f"""
                SELECT
                    p.*,
                    COUNT(t.id) AS total_tasks,
                    SUM(CASE WHEN t.status = 'done' THEN 1 ELSE 0 END) AS completed_tasks,
                    SUM(CASE WHEN t.status = 'blocked' THEN 1 ELSE 0 END) AS blocked_tasks
                FROM projects p
                LEFT JOIN tasks t ON t.project_id = p.id
                WHERE {where_clause}
                GROUP BY p.id
                ORDER BY {order_column} DESC, p.id DESC
                """
            ).fetchall()
        projects = []
        for row in rows:
            item = dict(row)
            item["total_tasks"] = int(item.get("total_tasks") or 0)
            item["completed_tasks"] = int(item.get("completed_tasks") or 0)
            item["blocked_tasks"] = int(item.get("blocked_tasks") or 0)
            projects.append(item)
        return projects

    def list_deleted_projects(self) -> list[dict]:
        return self.list_projects(deleted=True)

    def _load_project_detail(self, connection: sqlite3.Connection, project_id: int) -> dict:
        project_row = connection.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project_row or project_row["deleted_at"]:
            raise KeyError("project not found")
        project = dict(project_row)

        task_rows = connection.execute(
            """
            SELECT *
            FROM tasks
            WHERE project_id = ?
            ORDER BY sort_order ASC, id ASC
            """,
            (project_id,),
        ).fetchall()
        tasks = [row_to_dict(row) for row in task_rows]

        dependency_rows = connection.execute(
            """
            SELECT td.task_id, td.depends_on_task_id, t.title AS depends_on_title
            FROM task_dependencies td
            JOIN tasks t ON t.id = td.depends_on_task_id
            WHERE td.task_id IN (SELECT id FROM tasks WHERE project_id = ?)
            """,
            (project_id,),
        ).fetchall()
        dependency_map: dict[int, list[int]] = defaultdict(list)
        dependency_title_map: dict[int, list[str]] = defaultdict(list)
        for row in dependency_rows:
            dependency_map[int(row["task_id"])].append(int(row["depends_on_task_id"]))
            dependency_title_map[int(row["task_id"])].append(row["depends_on_title"])

        for task in tasks:
            task["dependency_ids"] = dependency_map.get(task["id"], [])
            task["dependency_titles"] = dependency_title_map.get(task["id"], [])

        stats = self._build_stats(project, tasks)
        return {"project": project, "tasks": tasks, "stats": stats}

    def get_project_detail(self, project_id: int) -> dict:
        with self.connect() as connection:
            return self._load_project_detail(connection, project_id)

    def _build_stats(self, project: dict, tasks: list[dict]) -> dict:
        if not tasks:
            return {
                "total_tasks": 0,
                "completed_tasks": 0,
                "in_progress_tasks": 0,
                "blocked_tasks": 0,
                "estimated_hours": 0.0,
                "actual_hours": 0.0,
                "average_confidence": 0.0,
                "projected_finish": project.get("start_date"),
                "slip_days": 0,
                "health": "planned",
            }

        start_dates = [parse_date(task.get("start_date")) for task in tasks if task.get("start_date")]
        end_dates = [parse_date(task.get("end_date")) for task in tasks if task.get("end_date")]
        projected_finish = max(end_dates) if end_dates else parse_date(project.get("start_date"))
        due_date = parse_date(project.get("due_date"))
        slip_days = 0
        if projected_finish and due_date and projected_finish > due_date:
            slip_days = (projected_finish - due_date).days

        stats = {
            "total_tasks": len(tasks),
            "completed_tasks": sum(1 for task in tasks if task["status"] == "done"),
            "in_progress_tasks": sum(1 for task in tasks if task["status"] == "in_progress"),
            "blocked_tasks": sum(1 for task in tasks if task["status"] == "blocked"),
            "estimated_hours": round(sum(task["estimate_hours"] for task in tasks), 1),
            "actual_hours": round(sum(task["actual_hours"] for task in tasks), 1),
            "average_confidence": round(sum(task["confidence"] for task in tasks) / len(tasks), 2),
            "schedule_start": format_date(min(start_dates)) if start_dates else project.get("start_date"),
            "projected_finish": format_date(projected_finish) if projected_finish else None,
            "slip_days": slip_days,
        }
        if stats["blocked_tasks"] > 0 or slip_days > 2:
            stats["health"] = "at_risk"
        elif stats["completed_tasks"] == len(tasks):
            stats["health"] = "done"
        elif stats["in_progress_tasks"] > 0:
            stats["health"] = "active"
        else:
            stats["health"] = "planned"
        return stats

    def create_project(self, payload: dict, use_suggestions: bool = False) -> dict:
        with self.connect() as connection:
            start_date = payload.get("start_date", today_str())
            due_date = payload.get("due_date")
            description = str(payload.get("description", "") or "").strip()
            if use_suggestions and not description:
                raise ValueError("\u667a\u80fd\u521b\u5efa\u9700\u8981\u9879\u76ee\u63cf\u8ff0\u6216\u8def\u7ebf\u56fe\u6587\u672c")
            category = payload.get("category") or infer_category(payload["name"], description)
            analysis = {
                "source": "manual",
                "model": "",
                "note": "",
            }
            suggestions: list[dict] = []
            preserve_schedule = False
            if use_suggestions:
                suggestion_result = build_task_suggestions(
                    payload["name"],
                    description,
                    start_date,
                    due_date,
                    self.planner,
                )
                category = payload.get("category") or suggestion_result["category"]
                suggestions = suggestion_result["tasks"]
                analysis = {
                    "source": suggestion_result["source"],
                    "model": suggestion_result.get("model", ""),
                    "note": suggestion_result.get("note", ""),
                }
                description = suggestion_result.get("project_description", description)
                start_date = suggestion_result.get("start_date", start_date)
                due_date = suggestion_result.get("due_date", due_date)
                preserve_schedule = bool(suggestion_result.get("preserve_schedule"))

            project_id = self._insert_project(
                connection,
                {
                    "name": payload["name"],
                    "description": description,
                    "category": category,
                    "status": payload.get("status", "planned"),
                    "start_date": start_date,
                    "due_date": due_date,
                    "analysis_source": analysis["source"],
                    "analysis_model": analysis["model"],
                    "analysis_note": analysis["note"],
                },
            )
            if use_suggestions:
                self._insert_tasks_for_project(connection, project_id, suggestions)
                if not preserve_schedule:
                    self._recalculate_project(connection, project_id)
            self._create_project_snapshot(connection, project_id, "create_project", "项目创建后自动快照")
            detail = self._load_project_detail(connection, project_id)
            detail["analysis"] = analysis
            detail["imported_task_count"] = len(suggestions)
            return detail

    def update_project(self, project_id: int, payload: dict) -> dict:
        with self.connect() as connection:
            current = connection.execute("SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL", (project_id,)).fetchone()
            if not current:
                raise KeyError("project not found")
            merged = dict(current)
            merged.update(payload)
            merged["updated_at"] = now_iso()
            if not merged.get("category"):
                merged["category"] = infer_category(merged["name"], merged.get("description", ""))
            connection.execute(
                """
                UPDATE projects
                SET name = ?, description = ?, category = ?, status = ?, start_date = ?, due_date = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    merged["name"],
                    merged.get("description", ""),
                    merged.get("category", "general"),
                    merged.get("status", "planned"),
                    merged.get("start_date", today_str()),
                    merged.get("due_date"),
                    merged["updated_at"],
                    project_id,
                ),
            )
            self._create_project_snapshot(connection, project_id, "update_project", "项目信息更新后自动快照")
            return self._load_project_detail(connection, project_id)

    def delete_project(self, project_id: int) -> None:
        with self.connect() as connection:
            cursor = connection.execute(
                "UPDATE projects SET deleted_at = ?, updated_at = ? WHERE id = ? AND deleted_at IS NULL",
                (now_iso(), now_iso(), project_id),
            )
            if cursor.rowcount == 0:
                raise KeyError("project not found")

    def restore_project(self, project_id: int) -> dict:
        with self.connect() as connection:
            cursor = connection.execute(
                "UPDATE projects SET deleted_at = NULL, updated_at = ? WHERE id = ? AND deleted_at IS NOT NULL",
                (now_iso(), project_id),
            )
            if cursor.rowcount == 0:
                raise KeyError("project not found")
            self._create_project_snapshot(connection, project_id, "restore_project", "恢复项目后自动快照")
            return self._load_project_detail(connection, project_id)

    def create_task(self, payload: dict) -> dict:
        with self.connect() as connection:
            project = connection.execute("SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL", (payload["project_id"],)).fetchone()
            if not project:
                raise KeyError("project not found")

            dependency_ids = normalize_dependencies(payload.get("dependency_ids"))
            title = payload["title"]
            description = payload.get("description", "")
            complexity = parse_int(payload.get("complexity"), 3)
            estimate_hours = parse_float(payload.get("estimate_hours"), 0.0)
            confidence = parse_float(payload.get("confidence"), 0.0)
            estimate_basis = payload.get("estimate_basis", "")
            if estimate_hours <= 0:
                estimate_hours, confidence, estimate_basis = estimate_task_hours(
                    title,
                    description,
                    complexity,
                    len(dependency_ids),
                    project["category"],
                )

            start_date = payload.get("start_date") or project["start_date"]
            end_date = payload.get("end_date") or add_days(start_date, task_duration_days(estimate_hours) - 1)
            timestamp = now_iso()
            cursor = connection.execute(
                """
                INSERT INTO tasks (
                    project_id, parent_id, title, description, status, owner, priority, complexity,
                    estimate_hours, actual_hours, progress, start_date, end_date, confidence,
                    estimate_basis, notes, auto_generated, sort_order, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload["project_id"],
                    payload.get("parent_id"),
                    title,
                    description,
                    payload.get("status", "planned"),
                    payload.get("owner", ""),
                    parse_int(payload.get("priority"), 2),
                    complexity,
                    estimate_hours,
                    parse_float(payload.get("actual_hours"), 0.0),
                    normalize_progress(payload.get("status", "planned"), payload.get("progress")),
                    start_date,
                    end_date,
                    confidence or 0.56,
                    estimate_basis or "创建任务",
                    payload.get("notes", ""),
                    parse_int(payload.get("auto_generated"), 0),
                    parse_int(payload.get("sort_order"), 999),
                    timestamp,
                    timestamp,
                ),
            )
            task_id = int(cursor.lastrowid)
            self._replace_dependencies(connection, task_id, dependency_ids)
            self._insert_estimate_history(connection, task_id, estimate_hours, confidence or 0.56, estimate_basis or "创建任务")

            if payload.get("auto_schedule"):
                self._recalculate_project(connection, payload["project_id"])
            self._create_project_snapshot(connection, payload["project_id"], "create_task", "新增任务后自动快照")
            return self._load_project_detail(connection, payload["project_id"])

    def update_task(self, task_id: int, payload: dict) -> dict:
        with self.connect() as connection:
            current = connection.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
            if not current:
                raise KeyError("task not found")
            merged = dict(current)
            merged.update(payload)
            merged["priority"] = parse_int(merged.get("priority"), 2)
            merged["complexity"] = parse_int(merged.get("complexity"), 3)
            merged["estimate_hours"] = parse_float(merged.get("estimate_hours"), 6.0)
            merged["actual_hours"] = parse_float(merged.get("actual_hours"), 0.0)
            merged["progress"] = normalize_progress(merged.get("status", "planned"), merged.get("progress"))
            merged["confidence"] = parse_float(merged.get("confidence"), 0.56)
            merged["updated_at"] = now_iso()
            if not merged.get("start_date"):
                project_start = connection.execute(
                    "SELECT start_date FROM projects WHERE id = ?", (merged["project_id"],)
                ).fetchone()[0]
                merged["start_date"] = project_start
            if not merged.get("end_date"):
                merged["end_date"] = add_days(merged["start_date"], task_duration_days(merged["estimate_hours"]) - 1)

            connection.execute(
                """
                UPDATE tasks
                SET parent_id = ?, title = ?, description = ?, status = ?, owner = ?, priority = ?, complexity = ?,
                    estimate_hours = ?, actual_hours = ?, progress = ?, start_date = ?, end_date = ?, confidence = ?,
                    estimate_basis = ?, notes = ?, auto_generated = ?, sort_order = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    merged.get("parent_id"),
                    merged["title"],
                    merged.get("description", ""),
                    merged.get("status", "planned"),
                    merged.get("owner", ""),
                    merged["priority"],
                    merged["complexity"],
                    merged["estimate_hours"],
                    merged["actual_hours"],
                    merged["progress"],
                    merged.get("start_date"),
                    merged.get("end_date"),
                    merged["confidence"],
                    merged.get("estimate_basis", ""),
                    merged.get("notes", ""),
                    parse_int(merged.get("auto_generated"), 0),
                    parse_int(merged.get("sort_order"), 999),
                    merged["updated_at"],
                    task_id,
                ),
            )
            if "dependency_ids" in payload:
                self._replace_dependencies(connection, task_id, normalize_dependencies(payload.get("dependency_ids")))
            if "estimate_hours" in payload or "confidence" in payload or "estimate_basis" in payload:
                self._insert_estimate_history(
                    connection,
                    task_id,
                    merged["estimate_hours"],
                    merged["confidence"],
                    merged.get("estimate_basis", "更新任务"),
                )
            if payload.get("auto_schedule"):
                self._recalculate_project(connection, merged["project_id"])
            self._create_project_snapshot(connection, merged["project_id"], "update_task", "编辑任务后自动快照")
            return self._load_project_detail(connection, merged["project_id"])

    def delete_task(self, task_id: int) -> dict:
        with self.connect() as connection:
            row = connection.execute("SELECT project_id FROM tasks WHERE id = ?", (task_id,)).fetchone()
            if not row:
                raise KeyError("task not found")
            project_id = int(row["project_id"])
            connection.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
            self._create_project_snapshot(connection, project_id, "delete_task", "删除任务后自动快照")
            return self._load_project_detail(connection, project_id)

    def _recalculate_project(self, connection: sqlite3.Connection, project_id: int) -> dict:
        project_row = connection.execute("SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL", (project_id,)).fetchone()
        if not project_row:
            raise KeyError("project not found")
        project_start = parse_date(project_row["start_date"]) or date.today()
        tasks = connection.execute(
            "SELECT id, estimate_hours, sort_order FROM tasks WHERE project_id = ? ORDER BY sort_order, id",
            (project_id,),
        ).fetchall()
        task_map = {int(row["id"]): row for row in tasks}
        dependency_rows = connection.execute(
            """
            SELECT task_id, depends_on_task_id
            FROM task_dependencies
            WHERE task_id IN (SELECT id FROM tasks WHERE project_id = ?)
            """,
            (project_id,),
        ).fetchall()
        graph: dict[int, list[int]] = defaultdict(list)
        indegree = {task_id: 0 for task_id in task_map}
        reverse_dependencies: dict[int, list[int]] = defaultdict(list)
        for row in dependency_rows:
            task_id = int(row["task_id"])
            depends_on_id = int(row["depends_on_task_id"])
            graph[depends_on_id].append(task_id)
            indegree[task_id] = indegree.get(task_id, 0) + 1
            reverse_dependencies[task_id].append(depends_on_id)

        order: list[int] = []
        queue = deque(
            sorted(
                (task_id for task_id, value in indegree.items() if value == 0),
                key=lambda tid: (task_map[tid]["sort_order"], tid),
            )
        )
        while queue:
            current = queue.popleft()
            order.append(current)
            for neighbor in sorted(graph.get(current, []), key=lambda tid: (task_map[tid]["sort_order"], tid)):
                indegree[neighbor] -= 1
                if indegree[neighbor] == 0:
                    queue.append(neighbor)

        if len(order) != len(task_map):
            order = sorted(task_map, key=lambda tid: (task_map[tid]["sort_order"], tid))

        schedule_end: dict[int, date] = {}
        for task_id in order:
            dependency_end_dates = [schedule_end[dependency] for dependency in reverse_dependencies.get(task_id, []) if dependency in schedule_end]
            start_date = max([project_start, *dependency_end_dates], default=project_start)
            if dependency_end_dates:
                start_date += timedelta(days=1)
            duration_days = task_duration_days(task_map[task_id]["estimate_hours"])
            end_date = start_date + timedelta(days=duration_days - 1)
            connection.execute(
                "UPDATE tasks SET start_date = ?, end_date = ?, updated_at = ? WHERE id = ?",
                (format_date(start_date), format_date(end_date), now_iso(), task_id),
            )
            schedule_end[task_id] = end_date

        connection.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (now_iso(), project_id))
        return self._load_project_detail(connection, project_id)

    def recalculate_project(self, project_id: int) -> dict:
        with self.connect() as connection:
            self._recalculate_project(connection, project_id)
            self._create_project_snapshot(connection, project_id, "reschedule", "重算排期后自动快照")
            return self._load_project_detail(connection, project_id)

    def import_tasks(self, payload: dict) -> dict:
        file_format = str(payload.get("format", "")).lower()
        if file_format not in {"csv", "json", "xlsx"}:
            raise ValueError("unsupported format")
        content = base64.b64decode(payload["content_base64"])
        rows = self._read_import_rows(content, file_format)
        if not rows:
            raise ValueError("empty import")

        with self.connect() as connection:
            project_id = parse_int(payload.get("project_id"))
            if project_id > 0:
                project_row = connection.execute("SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL", (project_id,)).fetchone()
                if not project_row:
                    raise KeyError("project not found")
            else:
                project_name = payload.get("project_name") or Path(str(payload.get("file_name", "导入任务"))).stem
                project_description = f"从 {payload.get('file_name', '文件')} 导入的任务。"
                project_start = payload.get("start_date", today_str())
                project_due = payload.get("due_date")
                project_id = self._insert_project(
                    connection,
                    {
                        "name": project_name,
                        "description": project_description,
                        "category": infer_category(project_name, project_description),
                        "status": "planned",
                        "start_date": project_start,
                        "due_date": project_due,
                    },
                )
                project_row = connection.execute("SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL", (project_id,)).fetchone()

            if payload.get("replace_existing"):
                connection.execute("DELETE FROM tasks WHERE project_id = ?", (project_id,))
                sort_offset = 0
            else:
                sort_offset = parse_int(
                    connection.execute(
                        "SELECT COALESCE(MAX(sort_order), -1) FROM tasks WHERE project_id = ?",
                        (project_id,),
                    ).fetchone()[0],
                    -1,
                ) + 1

            tasks, deferred_links = self._rows_to_tasks(rows, dict(project_row))
            created_ids = self._insert_tasks_for_project(connection, project_id, tasks, sort_offset=sort_offset)

            title_lookup = {
                tasks[index]["title"].strip().lower(): created_ids[index] for index in range(min(len(tasks), len(created_ids)))
            }
            for index, dependency_tokens, parent_token in deferred_links:
                if index > len(created_ids):
                    continue
                task_id = created_ids[index - 1]
                resolved_dependencies = []
                for token in dependency_tokens:
                    dependency_id = title_lookup.get(token.strip().lower(), parse_int(token))
                    if dependency_id:
                        resolved_dependencies.append(dependency_id)
                self._replace_dependencies(connection, task_id, resolved_dependencies)
                parent_id = title_lookup.get((parent_token or "").strip().lower(), parse_int(parent_token))
                if parent_id and parent_id != task_id:
                    connection.execute("UPDATE tasks SET parent_id = ? WHERE id = ?", (parent_id, task_id))

            connection.execute(
                """
                INSERT INTO import_export_log (project_id, action, file_format, row_count, created_at)
                VALUES (?, 'import', ?, ?, ?)
                """,
                (project_id, file_format, len(rows), now_iso()),
            )
            self._recalculate_project(connection, project_id)
            self._create_project_snapshot(connection, project_id, "import_tasks", "导入任务后自动快照")
            return self._load_project_detail(connection, project_id)

    def smart_import_tasks(self, project_id: int, payload: dict) -> dict:
        text_input = str(payload.get("description") or payload.get("text") or "").strip()
        if not text_input:
            raise ValueError("\u8bf7\u5148\u63d0\u4f9b\u8981\u89e3\u6790\u7684\u6587\u672c")

        with self.connect() as connection:
            project_row = connection.execute("SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL", (project_id,)).fetchone()
            if not project_row:
                raise KeyError("project not found")

            project = dict(project_row)
            replace_existing = parse_bool(payload.get("replace_existing"), False)
            suggestion_result = build_task_suggestions(
                project["name"],
                text_input,
                project.get("start_date") or today_str(),
                project.get("due_date"),
                self.planner,
            )
            tasks = suggestion_result.get("tasks") or []
            if not tasks:
                raise ValueError("\u672a\u80fd\u89e3\u6790\u51fa\u53ef\u5bfc\u5165\u7684\u4efb\u52a1")

            if replace_existing:
                connection.execute("DELETE FROM tasks WHERE project_id = ?", (project_id,))
                sort_offset = 0
                batch_start = project.get("start_date") or today_str()
            else:
                sort_offset = parse_int(
                    connection.execute(
                        "SELECT COALESCE(MAX(sort_order), -1) FROM tasks WHERE project_id = ?",
                        (project_id,),
                    ).fetchone()[0],
                    -1,
                ) + 1
                last_end = connection.execute(
                    "SELECT MAX(end_date) FROM tasks WHERE project_id = ?",
                    (project_id,),
                ).fetchone()[0]
                batch_start = add_days(last_end, 1) if last_end else (project.get("start_date") or today_str())

            if not suggestion_result.get("preserve_schedule"):
                schedule_task_batch(tasks, batch_start)

            created_ids = self._insert_tasks_for_project(connection, project_id, tasks, sort_offset=sort_offset)
            latest_end = max((parse_date(task.get("end_date")) for task in tasks if task.get("end_date")), default=None)
            project_description = str(project.get("description") or "").strip()
            if suggestion_result.get("project_description") and not project_description:
                connection.execute(
                    "UPDATE projects SET description = ?, updated_at = ? WHERE id = ?",
                    (suggestion_result.get("project_description", ""), now_iso(), project_id),
                )
                project["description"] = suggestion_result.get("project_description", "")
            current_due = parse_date(project.get("due_date"))
            if latest_end and (replace_existing or current_due is None or latest_end > current_due):
                connection.execute(
                    "UPDATE projects SET due_date = ?, updated_at = ? WHERE id = ?",
                    (format_date(latest_end), now_iso(), project_id),
                )

            connection.execute(
                """
                INSERT INTO import_export_log (project_id, action, file_format, row_count, created_at)
                VALUES (?, 'import', ?, ?, ?)
                """,
                (project_id, "smart_text", len(created_ids), now_iso()),
            )
            self._create_project_snapshot(connection, project_id, "smart_import", "智能追加任务后自动快照")
            detail = self._load_project_detail(connection, project_id)
            detail["analysis"] = {
                "source": suggestion_result.get("source", "rules"),
                "model": suggestion_result.get("model", ""),
                "note": suggestion_result.get("note", ""),
            }
            detail["imported_task_count"] = len(created_ids)
            detail["replace_existing"] = replace_existing
            return detail


    def meeting_update_progress(self, project_id: int, payload: dict) -> dict:
        meeting_text = str(payload.get("meeting_text") or payload.get("text") or "").strip()
        if not meeting_text:
            raise ValueError("请先提供本次会议纪要内容")
        if not self.planner:
            raise ValueError("未配置 SiliconFlow API Key，无法自动更新会议进度")

        auto_schedule = parse_bool(payload.get("auto_schedule"), False)
        with self.connect() as connection:
            project_row = connection.execute(
                "SELECT * FROM projects WHERE id = ? AND deleted_at IS NULL",
                (project_id,),
            ).fetchone()
            if not project_row:
                raise KeyError("project not found")

            task_rows = connection.execute(
                """
                SELECT *
                FROM tasks
                WHERE project_id = ?
                ORDER BY sort_order ASC, id ASC
                """,
                (project_id,),
            ).fetchall()
            tasks = [row_to_dict(row) for row in task_rows]
            if not tasks:
                raise ValueError("当前项目没有任务可更新")

            llm_result = self.planner.analyze_meeting_updates(dict(project_row), tasks, meeting_text)
            raw_updates = llm_result.get("updates") or []
            if not raw_updates:
                raise ValueError("未从会议内容中识别到可更新的任务进度")

            task_map = {int(task["id"]): task for task in tasks}
            applied_updates: list[dict] = []
            skipped_count = 0
            meeting_time = datetime.now().strftime("%Y-%m-%d %H:%M")

            for item in raw_updates:
                if not isinstance(item, dict):
                    skipped_count += 1
                    continue

                task_id = parse_int(item.get("task_id"), 0)
                if task_id <= 0 or task_id not in task_map:
                    skipped_count += 1
                    continue

                current = task_map[task_id]
                current_progress = int(clamp(parse_int(current.get("progress"), 0), 0, 100))
                next_progress = int(clamp(parse_int(item.get("progress"), current_progress), 0, 100))

                next_status = str(item.get("status") or "").strip()
                if next_status not in {"planned", "in_progress", "blocked", "done"}:
                    next_status = infer_status_from_progress(next_progress, str(current.get("status") or "planned"))

                reason = str(item.get("reason") or "").strip()
                if len(reason) > 180:
                    reason = reason[:180].rstrip("，,；; ") + "…"

                if current_progress == next_progress and str(current.get("status") or "") == next_status:
                    skipped_count += 1
                    continue

                note_line = f"[会议更新 {meeting_time}] 进度 {current_progress}% -> {next_progress}%"
                if reason:
                    note_line += f"；依据：{reason}"
                original_notes = str(current.get("notes") or "").strip()
                next_notes = f"{original_notes}\n{note_line}".strip() if original_notes else note_line

                connection.execute(
                    "UPDATE tasks SET progress = ?, status = ?, notes = ?, updated_at = ? WHERE id = ?",
                    (next_progress, next_status, next_notes, now_iso(), task_id),
                )

                current["progress"] = next_progress
                current["status"] = next_status
                current["notes"] = next_notes
                applied_updates.append(
                    {
                        "task_id": task_id,
                        "title": current.get("title", ""),
                        "progress": next_progress,
                        "status": next_status,
                        "reason": reason,
                    }
                )

            if not applied_updates:
                raise ValueError("识别到的更新与当前任务状态一致，无需写入")

            summary = str(llm_result.get("summary") or "").strip()
            if not summary:
                summary = f"会议纪要自动更新了 {len(applied_updates)} 个任务进度"

            connection.execute(
                """
                UPDATE projects
                SET analysis_source = ?, analysis_model = ?, analysis_note = ?, updated_at = ?
                WHERE id = ?
                """,
                ("llm_meeting", self.planner.model, summary, now_iso(), project_id),
            )
            connection.execute(
                """
                INSERT INTO import_export_log (project_id, action, file_format, row_count, created_at)
                VALUES (?, 'meeting_update', ?, ?, ?)
                """,
                (project_id, "meeting_text", len(applied_updates), now_iso()),
            )

            if auto_schedule:
                detail = self._recalculate_project(connection, project_id)
            else:
                detail = self._load_project_detail(connection, project_id)

            self._create_project_snapshot(connection, project_id, "meeting_update", summary)
            detail = self._load_project_detail(connection, project_id)
            detail["meeting_update"] = {
                "source": "llm",
                "model": self.planner.model,
                "summary": summary,
                "updated_count": len(applied_updates),
                "skipped_count": skipped_count,
                "updates": applied_updates,
                "auto_schedule": auto_schedule,
            }
            return detail
    def _read_import_rows(self, content: bytes, file_format: str) -> list[dict]:
        if file_format == "csv":
            text = content.decode("utf-8-sig")
            return [dict(row) for row in csv.DictReader(io.StringIO(text))]
        if file_format == "json":
            payload = json.loads(content.decode("utf-8"))
            if isinstance(payload, dict) and "tasks" in payload:
                return [dict(row) for row in payload["tasks"]]
            if isinstance(payload, list):
                return [dict(row) for row in payload]
            raise ValueError("invalid json structure")
        if file_format == "xlsx":
            if not load_workbook:
                raise ValueError("xlsx not available")
            workbook = load_workbook(io.BytesIO(content), data_only=True)
            sheet = workbook["任务"] if "任务" in workbook.sheetnames else workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            parsed_rows = []
            for values in rows[1:]:
                item = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    item[header] = values[index] if index < len(values) else None
                if any(value not in ("", None) for value in item.values()):
                    parsed_rows.append(item)
            return parsed_rows
        raise ValueError("unsupported format")

    def _rows_to_tasks(self, rows: list[dict], project: dict) -> tuple[list[dict], list[tuple[int, list[str], str | None]]]:
        synonym_map = {
            "title": {"title", "task", "name", "任务", "任务名"},
            "description": {"description", "desc", "描述"},
            "status": {"status", "状态"},
            "owner": {"owner", "负责人", "assignee"},
            "priority": {"priority", "优先级"},
            "complexity": {"complexity", "复杂度"},
            "estimate_hours": {"estimate_hours", "estimate", "预估工时", "工时"},
            "actual_hours": {"actual_hours", "actual", "实际工时"},
            "start_date": {"start_date", "start", "开始日期"},
            "end_date": {"end_date", "end", "due", "结束日期", "截止日期"},
            "depends_on": {"depends_on", "dependency", "依赖", "前置任务"},
            "parent": {"parent", "parent_id", "父任务"},
            "notes": {"notes", "remark", "备注"},
        }

        lowered_synonyms = {
            target: {item.lower() for item in values} for target, values in synonym_map.items()
        }

        def pick(row: dict, target: str):
            for key, value in row.items():
                normalized = str(key).strip().lower()
                if normalized in lowered_synonyms[target]:
                    return value
            return None

        tasks = []
        deferred_links = []
        for index, row in enumerate(rows):
            title = str(pick(row, "title") or "").strip()
            if not title:
                continue
            description = str(pick(row, "description") or "").strip()
            dependency_text = str(pick(row, "depends_on") or "").replace("，", ",")
            parent_text = str(pick(row, "parent") or "").strip() or None
            dependencies = [token.strip() for token in dependency_text.split(",") if token.strip()]
            complexity = parse_int(pick(row, "complexity"), 3)
            estimate_hours = parse_float(pick(row, "estimate_hours"), 0.0)
            confidence = 0.56
            estimate_basis = "导入任务"
            if estimate_hours <= 0:
                estimate_hours, confidence, estimate_basis = estimate_task_hours(
                    title,
                    description,
                    complexity,
                    len(dependencies),
                    project["category"],
                )
            task = {
                "key": f"import-{index}",
                "title": title,
                "description": description,
                "status": str(pick(row, "status") or "planned"),
                "owner": str(pick(row, "owner") or ""),
                "priority": parse_int(pick(row, "priority"), 2),
                "complexity": complexity,
                "estimate_hours": estimate_hours,
                "actual_hours": parse_float(pick(row, "actual_hours"), 0.0),
                "progress": None,
                "start_date": format_date(parse_date(str(pick(row, "start_date") or ""))) or project["start_date"],
                "end_date": format_date(parse_date(str(pick(row, "end_date") or ""))) or project["start_date"],
                "dependency_keys": [],
                "confidence": confidence,
                "estimate_basis": estimate_basis,
                "notes": str(pick(row, "notes") or ""),
                "parent_ref": None,
                "auto_generated": 0,
            }
            tasks.append(task)
            deferred_links.append((len(tasks), dependencies, parent_text))

        return tasks, deferred_links

    def export_project(self, project_id: int, file_format: str) -> tuple[str, bytes, str]:
        with self.connect() as connection:
            detail = self._load_project_detail(connection, project_id)
            project = detail["project"]
            tasks = detail["tasks"]
            file_format = file_format.lower()
            file_name = safe_filename(project["name"])
            if file_format == "json":
                payload = json.dumps(detail, ensure_ascii=False, indent=2).encode("utf-8")
                mime = "application/json; charset=utf-8"
                extension = "json"
            elif file_format == "csv":
                buffer = io.StringIO()
                writer = csv.DictWriter(
                    buffer,
                    fieldnames=[
                        "id",
                        "title",
                        "status",
                        "owner",
                        "priority",
                        "complexity",
                        "estimate_hours",
                        "actual_hours",
                        "progress",
                        "start_date",
                        "end_date",
                        "dependency_ids",
                        "notes",
                    ],
                )
                writer.writeheader()
                for task in tasks:
                    writer.writerow(
                        {
                            "id": task["id"],
                            "title": task["title"],
                            "status": task["status"],
                            "owner": task["owner"],
                            "priority": task["priority"],
                            "complexity": task["complexity"],
                            "estimate_hours": task["estimate_hours"],
                            "actual_hours": task["actual_hours"],
                            "progress": task["progress"],
                            "start_date": task["start_date"],
                            "end_date": task["end_date"],
                            "dependency_ids": ",".join(str(item) for item in task["dependency_ids"]),
                            "notes": task["notes"],
                        }
                    )
                payload = buffer.getvalue().encode("utf-8-sig")
                mime = "text/csv; charset=utf-8"
                extension = "csv"
            elif file_format == "xlsx":
                if not Workbook or not PatternFill or not FormulaRule or not DataBarRule or not get_column_letter:
                    raise ValueError("xlsx not available")
                workbook = Workbook()
                build_excel_gantt_sheet(workbook, project, tasks)
                task_sheet = workbook.create_sheet("任务")
                build_excel_task_sheet(task_sheet, tasks)
                project_sheet = workbook.create_sheet("项目")
                build_excel_project_sheet(project_sheet, project)
                workbook.active = 0
                binary = io.BytesIO()
                workbook.save(binary)
                payload = binary.getvalue()
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                extension = "xlsx"
            else:
                raise ValueError("unsupported format")

            connection.execute(
                """
                INSERT INTO import_export_log (project_id, action, file_format, row_count, created_at)
                VALUES (?, 'export', ?, ?, ?)
                """,
                (project_id, file_format, len(tasks), now_iso()),
            )
            return f"{file_name}.{extension}", payload, mime

    def export_all_projects(self, file_format: str) -> tuple[str, bytes, str]:
        file_format = file_format.lower()
        if file_format != "xlsx":
            raise ValueError("unsupported format")
        if not Workbook or not PatternFill or not FormulaRule or not DataBarRule or not get_column_letter:
            raise ValueError("xlsx not available")

        with self.connect() as connection:
            project_rows = connection.execute(
                "SELECT id FROM projects WHERE deleted_at IS NULL ORDER BY COALESCE(start_date, ''), updated_at DESC, id DESC"
            ).fetchall()
            if not project_rows:
                raise ValueError("no projects")

            details = [self._load_project_detail(connection, int(row["id"])) for row in project_rows]
            workbook = Workbook()
            build_excel_all_projects_gantt_sheet(workbook, details)
            task_sheet = workbook.create_sheet("全部任务")
            build_excel_all_task_sheet(task_sheet, details)
            project_sheet = workbook.create_sheet("项目总表")
            build_excel_project_list_sheet(project_sheet, details)
            workbook.active = 0
            binary = io.BytesIO()
            workbook.save(binary)
            payload = binary.getvalue()
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            total_task_count = sum(len(detail["tasks"]) for detail in details)
            connection.execute(
                """
                INSERT INTO import_export_log (project_id, action, file_format, row_count, created_at)
                VALUES (?, 'export', ?, ?, ?)
                """,
                (None, "xlsx_all", total_task_count, now_iso()),
            )
            return "all_projects_gantt.xlsx", payload, mime

class AppServer(ThreadingHTTPServer):
    daemon_threads = True

    def __init__(self, server_address, handler_class, db: TaskDatabase):
        super().__init__(server_address, handler_class)
        self.db = db


class TaskGanttHandler(BaseHTTPRequestHandler):
    server_version = "TaskGantt/1.0"

    def do_GET(self) -> None:
        try:
            parsed = urlparse(self.path)
            if parsed.path.startswith("/api/"):
                self.handle_api_get(parsed)
            else:
                self.serve_static(parsed.path)
        except Exception as exc:  # pragma: no cover - defensive path
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_POST(self) -> None:
        self.handle_api_write("POST")

    def do_PUT(self) -> None:
        self.handle_api_write("PUT")

    def do_DELETE(self) -> None:
        self.handle_api_write("DELETE")

    def log_message(self, format: str, *args) -> None:  # pragma: no cover - quiet default logging
        print("%s - - [%s] %s" % (self.address_string(), self.log_date_time_string(), format % args))

    def read_json_body(self) -> dict:
        content_length = parse_int(self.headers.get("Content-Length"), 0)
        if content_length <= 0:
            return {}
        raw = self.rfile.read(content_length)
        return json.loads(raw.decode("utf-8"))

    def send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_file(self, payload: bytes, content_type: str, file_name: str) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(payload)))
        self.send_header("Content-Disposition", f'attachment; filename="{file_name}"')
        self.end_headers()
        self.wfile.write(payload)

    def serve_static(self, request_path: str) -> None:
        target = STATIC_DIR / "index.html" if request_path in ("", "/") else STATIC_DIR / request_path.lstrip("/")
        target = target.resolve()
        static_root = STATIC_DIR.resolve()
        if not target.exists() or (target != static_root / "index.html" and static_root not in target.parents):
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return
        content = target.read_bytes()
        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def handle_api_get(self, parsed) -> None:
        path = parsed.path
        query = parse_qs(parsed.query)
        db: TaskDatabase = self.server.db

        if path == "/api/bootstrap":
            projects = db.list_projects()
            deleted_projects = db.list_deleted_projects()
            selected_project_id = projects[0]["id"] if projects else None
            detail = db.get_project_detail(selected_project_id) if selected_project_id else None
            self.send_json({
                "projects": projects,
                "deleted_projects": deleted_projects,
                "selected_project_id": selected_project_id,
                "detail": detail,
            })
            return

        if path == "/api/projects":
            self.send_json({
                "projects": db.list_projects(),
                "deleted_projects": db.list_deleted_projects(),
            })
            return

        if path.startswith("/api/projects/"):
            parts = [part for part in path.strip("/").split("/") if part]
            if len(parts) == 3:
                project_id = parse_int(parts[2])
                self.send_json(db.get_project_detail(project_id))
                return
            if len(parts) == 4 and parts[3] == "snapshots":
                project_id = parse_int(parts[2])
                self.send_json({"snapshots": db.list_project_snapshots(project_id)})
                return

        if path == "/api/export-all":
            file_format = str(query.get("format", ["xlsx"])[0])
            file_name, payload, content_type = db.export_all_projects(file_format)
            self.send_file(payload, content_type, file_name)
            return

        if path == "/api/export":
            project_id = parse_int(query.get("project_id", [0])[0])
            file_format = str(query.get("format", ["json"])[0])
            file_name, payload, content_type = db.export_project(project_id, file_format)
            self.send_file(payload, content_type, file_name)
            return

        self.send_json({"error": "unknown endpoint"}, HTTPStatus.NOT_FOUND)

    def handle_api_write(self, method: str) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        db: TaskDatabase = self.server.db
        body = self.read_json_body()
        try:
            if method == "POST" and path == "/api/projects":
                detail = db.create_project(body, use_suggestions=bool(body.get("use_suggestions")))
                self.send_json(detail, HTTPStatus.CREATED)
                return

            if method == "PUT" and path.startswith("/api/projects/"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 3:
                    project_id = parse_int(parts[2])
                    self.send_json(db.update_project(project_id, body))
                    return

            if method == "DELETE" and path.startswith("/api/projects/"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 3:
                    project_id = parse_int(parts[2])
                    db.delete_project(project_id)
                    self.send_json({"ok": True})
                    return

            if method == "POST" and path.startswith("/api/projects/") and path.endswith("/restore"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 4:
                    project_id = parse_int(parts[2])
                    self.send_json(db.restore_project(project_id))
                    return

            if method == "POST" and path.startswith("/api/projects/") and path.endswith("/reschedule"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 4:
                    project_id = parse_int(parts[2])
                    self.send_json(db.recalculate_project(project_id))
                    return

            if method == "POST" and path.startswith("/api/projects/") and path.endswith("/snapshots"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 4 and parts[3] == "snapshots":
                    project_id = parse_int(parts[2])
                    self.send_json({"snapshots": db.list_project_snapshots(project_id)}, HTTPStatus.CREATED)
                    return

            if method == "POST" and path.startswith("/api/projects/") and path.endswith("/restore"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 6 and parts[3] == "snapshots" and parts[5] == "restore":
                    project_id = parse_int(parts[2])
                    snapshot_id = parse_int(parts[4])
                    self.send_json(db.restore_project_snapshot(project_id, snapshot_id), HTTPStatus.CREATED)
                    return
            if method == "POST" and path.startswith("/api/projects/") and path.endswith("/meeting-update"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 4:
                    project_id = parse_int(parts[2])
                    self.send_json(db.meeting_update_progress(project_id, body), HTTPStatus.CREATED)
                    return
            if method == "POST" and path.startswith("/api/projects/") and path.endswith("/smart-import"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 4:
                    project_id = parse_int(parts[2])
                    self.send_json(db.smart_import_tasks(project_id, body), HTTPStatus.CREATED)
                    return

            if method == "POST" and path == "/api/tasks":
                detail = db.create_task(body)
                self.send_json(detail, HTTPStatus.CREATED)
                return

            if method == "PUT" and path.startswith("/api/tasks/"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 3:
                    task_id = parse_int(parts[2])
                    self.send_json(db.update_task(task_id, body))
                    return

            if method == "DELETE" and path.startswith("/api/tasks/"):
                parts = [part for part in path.strip("/").split("/") if part]
                if len(parts) == 3:
                    task_id = parse_int(parts[2])
                    self.send_json(db.delete_task(task_id))
                    return

            if method == "POST" and path == "/api/import":
                detail = db.import_tasks(body)
                self.send_json(detail, HTTPStatus.CREATED)
                return
        except KeyError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            return
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        except sqlite3.IntegrityError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return

        self.send_json({"error": "unknown endpoint"}, HTTPStatus.NOT_FOUND)


def create_server(host: str | None = None, port: int | None = None) -> AppServer:
    server_host = host or os.environ.get("TASK_GANTT_HOST", "127.0.0.1")
    server_port = port if port is not None else int(os.environ.get("TASK_GANTT_PORT", os.environ.get("PORT", "8010")))
    planner = SiliconFlowPlanner.from_env()
    db = TaskDatabase(DB_PATH, planner)
    return AppServer((server_host, server_port), TaskGanttHandler, db)


def main() -> None:
    server = create_server()
    planner = server.db.planner
    llm_text = f"LLM: {planner.model}" if planner else "LLM: disabled (.env)"
    print(f"Task Gantt chart running at http://{server.server_address[0]}:{server.server_address[1]} - {llm_text}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()







































